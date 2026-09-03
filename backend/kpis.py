"""Dashboard de KPIs de personal -- dos fuentes de datos distintas, a
propósito (pedido explícito del usuario):

1. Plantilla ACTIVA y horas contratadas: el informe que exporta GO
   (importación puntual, ver import_excel/kpi_empleados) -- es una foto de
   "quién está en nómina ahora mismo", no cambia sola.
2. Bajas (para rotación mensual/anual y % NSPP): la tabla EN VIVO de
   Entrevista de Salida (entrevistas_salidas), que crece cada vez que RRHH
   da de alta una salida ahí -- así el dashboard se mantiene al día solo,
   sin depender de reimportar el Excel cada vez que se va alguien.

% de promoción interna: NO se calcula todavía -- no hay ninguna fuente de
datos con el historial de cambios de puesto (el usuario lo confirmó: "por
ahora no lo tenemos"). Se deja como placeholder en el frontend en vez de
inventar un número."""
import calendar
import datetime
import io
import re
import sqlite3

import xlrd
from openpyxl import load_workbook

from db import get_connection

# Los dos sistemas nombran los mismos centros de forma distinta (el Excel de
# GO trae "GO - TIENDA CALEIDO", Entrevista de Salida usa "Caleido") --
# normalizamos el nombre de GO a la forma corta al importar, para poder
# cruzar plantilla (Excel) con bajas (Entrevista de Salida) por el mismo
# nombre de centro. Confirmado 1 a 1 contra los centros reales en producción.
CENTROS_GO_A_CORTO = {
    "GO - FABRICA PARQUE SUR": "ParqueSur Fabrica",
    "GO - TIENDA PARQUE SUR": "ParqueSur Tienda",
    "GO - ADMIN CENTRAL": "Oficina Central",
    "GO - TIENDA LA GAVIA": "La Gavia",
    "GO - TIENDA PLENILUNIO": "Plenilunio",
    "GO - TIENDA PRINCESA": "Princesa",
    "GO - TIENDA CALEIDO": "Caleido",
    "GO - TIENDA GRANPLAZA2": "Gran Plaza 2",
}

HORAS_JORNADA_COMPLETA = 40  # pedido explícito del usuario

# Administración central ("GO - ADMIN CENTRAL" / "Oficina Central") no cuenta
# para estos KPIs -- son de tienda/fábrica, no de oficina. Se descarta tanto
# la plantilla activa como las bajas de ese centro antes de calcular nada.
CENTROS_EXCLUIDOS = {"Oficina Central"}

# Motivos SEPE de "no superó el periodo de prueba" -- por iniciativa de la
# empresa o del propio trabajador cuentan igual para este KPI (NSPP no
# distingue quién lo decidió, solo que la baja fue en periodo de prueba).
_MOTIVO_NSPP = "periodo de prueba"


def _normaliza(s):
    s = (s or "").strip().lower()
    for a, b in (("á", "a"), ("é", "e"), ("í", "i"), ("ó", "o"), ("ú", "u"), ("ü", "u"), ("ñ", "n")):
        s = s.replace(a, b)
    return s


_ALIAS_COLUMNAS = {
    "nombre del centro": "centro",
    "codigo empleado": "codigo_empleado",
    "nombre completo": "nombre",
    "fecha antiguedad": "fecha_antiguedad",
    "posicion/puesto de trabajo": "puesto",
    "porcentaje jornada": "porcentaje_jornada",
    "fecha de baja en compania": "fecha_baja",
    "motivo de baja de la compania": "motivo_baja",
}


def ensure_kpis_tables():
    conn = get_connection()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS kpi_empleados (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo_empleado TEXT NOT NULL UNIQUE,
            centro TEXT,
            nombre TEXT,
            fecha_antiguedad TEXT,
            puesto TEXT,
            porcentaje_jornada REAL,
            fecha_baja TEXT,
            motivo_baja TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS kpi_importaciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            archivo_nombre TEXT,
            importado_por TEXT,
            importado_en TEXT NOT NULL DEFAULT (datetime('now')),
            filas INTEGER
        )
    """)
    conn.commit()
    conn.close()


def _fecha_a_iso(valor):
    if valor is None:
        return None
    if isinstance(valor, (datetime.datetime, datetime.date)):
        return valor.strftime("%Y-%m-%d")
    texto = str(valor).strip()
    if not texto or texto.lower() == "nan":
        return None
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", texto)
    if m:
        d, mth, y = m.groups()
        try:
            return datetime.date(int(y), int(mth), int(d)).isoformat()
        except ValueError:
            return None
    return None


def _numero(valor):
    if valor is None:
        return None
    texto = str(valor).strip().replace(",", ".")
    if not texto or texto.lower() == "nan":
        return None
    try:
        return float(texto)
    except ValueError:
        return None


def _texto(valor):
    if valor is None:
        return None
    texto = str(valor).strip()
    return texto if texto and texto.lower() != "nan" else None


def _codigo_empleado(valor):
    texto = _texto(valor)
    if texto and texto.endswith(".0") and texto[:-2].isdigit():
        return texto[:-2]
    return texto


def _centro_normalizado(valor):
    texto = _texto(valor)
    if texto is None:
        return None
    return CENTROS_GO_A_CORTO.get(texto, texto)


def _leer_filas_xlsx(contenido):
    wb = load_workbook(io.BytesIO(contenido), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    filas = list(ws.iter_rows(values_only=True))
    wb.close()
    return filas


def _leer_filas_xls(contenido):
    wb = xlrd.open_workbook(file_contents=contenido)
    ws = wb.sheet_by_index(0)
    filas = []
    for r in range(ws.nrows):
        fila = []
        for c in range(ws.ncols):
            celda = ws.cell(r, c)
            if celda.ctype == xlrd.XL_CELL_DATE:
                fila.append(datetime.datetime(*xlrd.xldate_as_tuple(celda.value, wb.datemode)))
            else:
                fila.append(celda.value)
        filas.append(fila)
    return filas


def import_excel(contenido, nombre_archivo, subido_por):
    es_xls = nombre_archivo.lower().endswith(".xls") and not nombre_archivo.lower().endswith(".xlsx")
    try:
        filas = _leer_filas_xls(contenido) if es_xls else _leer_filas_xlsx(contenido)
    except Exception as exc:
        raise ValueError(f"No se pudo leer el archivo Excel: {exc}")
    if not filas:
        raise ValueError("El archivo está vacío")

    encabezado = [_normaliza(str(c)) for c in filas[0]]
    indice = {}
    for i, col in enumerate(encabezado):
        clave = _ALIAS_COLUMNAS.get(col)
        if clave:
            indice[clave] = i
    faltantes = [c for c in ("centro", "codigo_empleado", "nombre") if c not in indice]
    if faltantes:
        raise ValueError(f"Faltan columnas obligatorias en el Excel: {', '.join(faltantes)}")

    registros = []
    for fila in filas[1:]:
        if fila is None or all(v is None or str(v).strip() == "" for v in fila):
            continue
        def val(clave):
            i = indice.get(clave)
            return fila[i] if i is not None and i < len(fila) else None
        codigo = _codigo_empleado(val("codigo_empleado"))
        if not codigo:
            continue
        registros.append((
            codigo, _centro_normalizado(val("centro")), _texto(val("nombre")),
            _fecha_a_iso(val("fecha_antiguedad")), _texto(val("puesto")),
            _numero(val("porcentaje_jornada")), _fecha_a_iso(val("fecha_baja")),
            _texto(val("motivo_baja")),
        ))
    if not registros:
        raise ValueError("No se encontró ninguna fila de empleado válida (falta el código de empleado)")

    conn = get_connection()
    conn.execute("DELETE FROM kpi_empleados")
    conn.executemany("""
        INSERT INTO kpi_empleados
            (codigo_empleado, centro, nombre, fecha_antiguedad, puesto, porcentaje_jornada, fecha_baja, motivo_baja)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
    """, registros)
    conn.execute(
        "INSERT INTO kpi_importaciones (archivo_nombre, importado_por, filas) VALUES (?, ?, ?)",
        (nombre_archivo, subido_por, len(registros)),
    )
    conn.commit()
    conn.close()
    return {"filas": len(registros)}


def get_ultima_importacion():
    conn = get_connection()
    row = conn.execute("SELECT * FROM kpi_importaciones ORDER BY id DESC LIMIT 1").fetchone()
    conn.close()
    return dict(row) if row else None


def _tabla_existe(conn, nombre):
    return conn.execute(
        "SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (nombre,)
    ).fetchone() is not None


def _get_bajas(conn, empresa="kk"):
    """Todas las bajas registradas en Entrevista de Salida (cualquier oleada
    de esa empresa) -- fuente EN VIVO, crece sola cada vez que RRHH da de
    alta una salida ahí, sin depender de reimportar ningún Excel."""
    if not _tabla_existe(conn, "entrevistas_salidas") or not _tabla_existe(conn, "entrevistas_oleadas"):
        return []
    rows = conn.execute("""
        SELECT s.centro, s.fecha_baja, s.motivo
        FROM entrevistas_salidas s JOIN entrevistas_oleadas o ON o.id = s.oleada_id
        WHERE o.empresa = ? AND s.fecha_baja IS NOT NULL AND s.fecha_baja != ''
    """, (empresa,)).fetchall()
    return [dict(r) for r in rows]


def _es_nspp(motivo):
    return bool(motivo) and _MOTIVO_NSPP in motivo.lower()


def _es_nspp_empresario(motivo):
    """NSPP decidido por la empresa (motivo SEPE '07') -- el dato "malo" a
    propósito que comentó el usuario, provocado por la propia empresa para
    reducir horas, no una baja real de mercado. Se excluye para tener una
    lectura de rotación sin ese ruido."""
    if not motivo:
        return False
    m = motivo.lower()
    return _MOTIVO_NSPP in m and "instancia del empresario" in m


def marcar_baja_manual(codigo_empleado, fecha_baja, motivo_baja):
    """Corrige a mano un registro de kpi_empleados que el Excel de GO todavía
    trae como activo pero que Entrevista de Salida ya tiene registrado como
    baja (el Excel es una foto puntual, puede ir por detrás). Se pierde en
    la próxima importación -- si el Excel sigue sin traer la baja hay que
    volver a aplicarla."""
    conn = get_connection()
    cur = conn.execute(
        "UPDATE kpi_empleados SET fecha_baja = ?, motivo_baja = ? WHERE codigo_empleado = ?",
        (fecha_baja, motivo_baja, str(codigo_empleado)),
    )
    conn.commit()
    afectado = cur.rowcount
    conn.close()
    if not afectado:
        raise ValueError(f"No se encontró ningún empleado con código {codigo_empleado}")
    return {"ok": True}


def _mes(fecha_iso):
    """'2026-03-15' -> '2026-03' -- agrupa por mes calendario. Tolera fechas
    en otros formatos sueltos (dd/mm/aaaa) que a veces se cuelan al registrar
    una salida a mano."""
    if not fecha_iso:
        return None
    if re.match(r"^\d{4}-\d{2}", fecha_iso):
        return fecha_iso[:7]
    m = re.match(r"^\d{1,2}/(\d{1,2})/(\d{4})$", fecha_iso)
    if m:
        mth, y = m.groups()
        return f"{y}-{int(mth):02d}"
    return None


def _fin_de_mes(clave_mes):
    anio, mes = (int(x) for x in clave_mes.split("-"))
    ultimo_dia = calendar.monthrange(anio, mes)[1]
    return f"{anio:04d}-{mes:02d}-{ultimo_dia:02d}"


def _activos_a_fecha(empleados, fecha_corte):
    """Quiénes estaban de alta en una fecha concreta -- se reconstruye con
    la fecha de antigüedad (alta) y la fecha de baja que ya trae el Excel de
    plantilla para cada empleado, así se puede saber la plantilla/horas de
    cualquier mes pasado, no solo la de ahora mismo. Un empleado sin fecha
    de antigüedad se asume de alta desde siempre (dato que a veces falta en
    el Excel) en vez de excluirlo."""
    activos = []
    for e in empleados:
        antiguedad = e.get("fecha_antiguedad")
        baja = e.get("fecha_baja")
        if antiguedad and antiguedad > fecha_corte:
            continue
        if baja and baja <= fecha_corte:
            continue
        activos.append(e)
    return activos


def _headcount_y_horas_por_centro(empleados_activos):
    hc_centro = {}
    horas_centro = {}
    for e in empleados_activos:
        c = e["centro"] or "(sin centro)"
        hc_centro[c] = hc_centro.get(c, 0) + 1
        pct = e["porcentaje_jornada"]
        horas = (pct / 100 * HORAS_JORNADA_COMPLETA) if pct is not None else HORAS_JORNADA_COMPLETA
        horas_centro[c] = horas_centro.get(c, 0) + horas
    return (
        sorted(hc_centro.items(), key=lambda x: -x[1]),
        sorted([(c, round(h, 1)) for c, h in horas_centro.items()], key=lambda x: -x[1]),
    )


def compute_resumen():
    conn = get_connection()
    empleados = [dict(r) for r in conn.execute("SELECT * FROM kpi_empleados").fetchall()]
    bajas = _get_bajas(conn)
    conn.close()

    empleados = [e for e in empleados if e["centro"] not in CENTROS_EXCLUIDOS]
    bajas = [b for b in bajas if b["centro"] not in CENTROS_EXCLUIDOS]

    hoy = datetime.date.today()
    hoy_str = hoy.isoformat()

    activos = _activos_a_fecha(empleados, hoy_str)
    headcount_activo = len(activos)
    headcount_por_centro_lista, horas_por_centro_lista = _headcount_y_horas_por_centro(activos)
    headcount_por_centro = dict(headcount_por_centro_lista)

    # --- Serie mensual completa (todo el histórico de Entrevista de Salida) -
    # Se calcula para TODOS los meses con datos (no solo los últimos 12) para
    # que el frontend pueda filtrar por un rango de fechas cualquiera. Cada
    # mes recalcula también su propia plantilla/horas "a fecha de fin de ese
    # mes" (con fecha_antiguedad/fecha_baja) -- no es una foto fija, así el
    # gráfico de horas por centro también responde al filtro de fechas.
    bajas_por_mes = {}
    bajas_centro_mes = {}
    bajas_motivo_mes = {}
    for b in bajas:
        clave = _mes(b["fecha_baja"])
        if not clave:
            continue
        bajas_por_mes[clave] = bajas_por_mes.get(clave, 0) + 1
        centro = b["centro"] or "(sin centro)"
        bajas_centro_mes.setdefault(clave, {})
        bajas_centro_mes[clave][centro] = bajas_centro_mes[clave].get(centro, 0) + 1
        motivo = b["motivo"] or "(sin dato)"
        bajas_motivo_mes.setdefault(clave, {})
        bajas_motivo_mes[clave][motivo] = bajas_motivo_mes[clave].get(motivo, 0) + 1

    mes_actual = f"{hoy.year:04d}-{hoy.month:02d}"
    claves_con_datos = sorted(bajas_por_mes.keys())
    inicio = min(claves_con_datos[0], mes_actual) if claves_con_datos else mes_actual
    fin = max(claves_con_datos[-1], mes_actual) if claves_con_datos else mes_actual

    meses_disponibles = []
    anio_i, mes_i = (int(x) for x in inicio.split("-"))
    anio_f, mes_f = (int(x) for x in fin.split("-"))
    while (anio_i, mes_i) <= (anio_f, mes_f):
        meses_disponibles.append(f"{anio_i:04d}-{mes_i:02d}")
        mes_i += 1
        if mes_i > 12:
            mes_i = 1
            anio_i += 1

    # NOTA: solo se reconstruye el TOTAL de plantilla activa por mes (para el
    # % de rotación), no el desglose por centro -- el Excel solo trae la
    # fecha de antigüedad en la EMPRESA y el centro ACTUAL, no una fecha de
    # alta por centro, así que no hay forma fiable de saber en qué tienda
    # trabajaba alguien en el pasado si se trasladó entre centros (p.ej. una
    # tienda que abrió después mostraría gente "trabajando" ahí antes de
    # existir). Para eso haría falta un histórico de movimientos internos
    # (puesto Y centro) que hoy no existe -- mismo caso que % promoción
    # interna, pendiente de una fuente de datos futura.
    serie_mensual = {}
    for clave in meses_disponibles:
        n = bajas_por_mes.get(clave, 0)
        headcount_mes = len(_activos_a_fecha(empleados, _fin_de_mes(clave)))
        serie_mensual[clave] = {
            "bajas": n,
            "pct": round(n / headcount_mes * 100, 1) if headcount_mes else 0,
            "por_centro": sorted(bajas_centro_mes.get(clave, {}).items(), key=lambda x: -x[1]),
            "por_motivo": sorted(bajas_motivo_mes.get(clave, {}).items(), key=lambda x: -x[1]),
            "headcount_activo": headcount_mes,
        }
    anios_disponibles = sorted({m[:4] for m in meses_disponibles}) or [str(hoy.year)]

    # --- Acumulado anual (año natural, desde el 1 de enero) ----------------
    inicio_anio = f"{hoy.year}-01-01"
    bajas_ytd = [b for b in bajas if b["fecha_baja"] and b["fecha_baja"] >= inicio_anio]
    acumulado_anual_pct = round(len(bajas_ytd) / headcount_activo * 100, 1) if headcount_activo else 0

    # --- % NSPP (sobre las bajas del año en curso) --------------------------
    nspp_ytd = [b for b in bajas_ytd if _es_nspp(b["motivo"])]
    nspp_pct = round(len(nspp_ytd) / len(bajas_ytd) * 100, 1) if bajas_ytd else 0

    # --- Rotación anual quitando los NSPP decididos por la empresa ---------
    # Mismo acumulado anual, pero sin los "07 Cese en periodo de prueba a
    # instancia del empresario" -- ese es el dato "malo a propósito" que
    # comentó el usuario (provocado por la empresa para bajar horas), así se
    # puede ver la rotación real sin ese ruido.
    bajas_ytd_sin_empresario = [b for b in bajas_ytd if not _es_nspp_empresario(b["motivo"])]
    rotacion_sin_nspp_empresario_pct = (
        round(len(bajas_ytd_sin_empresario) / headcount_activo * 100, 1) if headcount_activo else 0
    )

    return {
        "headcount_activo": headcount_activo,
        "horas_contratadas_totales": round(sum(h for _, h in horas_por_centro_lista), 1),
        "horas_jornada_completa": HORAS_JORNADA_COMPLETA,
        "mes_actual": mes_actual,
        "serie_mensual": serie_mensual,
        "meses_disponibles": meses_disponibles,
        "anios_disponibles": anios_disponibles,
        "headcount_por_centro": headcount_por_centro_lista,
        "acumulado_anual_pct": acumulado_anual_pct,
        "bajas_ytd": len(bajas_ytd),
        "nspp_pct": nspp_pct,
        "nspp_ytd": len(nspp_ytd),
        "rotacion_sin_nspp_empresario_pct": rotacion_sin_nspp_empresario_pct,
        "bajas_sin_nspp_empresario_ytd": len(bajas_ytd_sin_empresario),
        "horas_por_centro": horas_por_centro_lista,
        "sin_datos_plantilla": headcount_activo == 0,
        "sin_datos_bajas": len(bajas) == 0,
    }


ensure_kpis_tables()
