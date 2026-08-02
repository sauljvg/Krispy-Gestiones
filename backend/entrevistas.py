"""Motor de cálculo de "Entrevista de Salida" — inspirado en el script de
Apps Script del usuario, pero detectando bloques y preguntas por el
CONTENIDO del encabezado en vez de una lista fija de preguntas por empresa.

Los formularios de Krispy Kreme y Saona usan la misma convención de
encabezado para las preguntas de escala: "Indica tu nivel de acuerdo con
las siguientes afirmaciones sobre <bloque>: [<pregunta>]" — el bloque es el
texto antes de los dos puntos, la pregunta el texto entre corchetes. Así
funciona igual para las dos empresas aunque redacten "tu llegada a Saona"
o "tu llegada a Krispy Kreme" de forma distinta.
"""
import hashlib
import io
import json
import re
from datetime import datetime

from openpyxl import load_workbook

from db import get_connection

METADATA_HINTS = {
    "marca temporal": "hora_inicio",
    "hora de inicio": "hora_inicio",
    "hora de finalizacion": "hora_fin",
    "correo electronico": "correo",
    "nombre": "nombre",
    "puesto de trabajo": "puesto",
}

_RESPUESTAS_SIN_VALOR = {".", "..", "…", "-", "sin comentarios", "sin comentario"}

# El campo "Centro de Trabajo" del formulario es texto libre, y a veces
# alguien escribe una variante del nombre real de la tienda (p.ej. "Princesa
# 7" en vez de "Princesa", el nombre que usa el resto de la app). Sin esto,
# esa persona cuenta como si fuera OTRO centro distinto.
_CENTRO_ALIASES = {
    "princesa 7": "Princesa",
    # _normaliza_header ya quita acentos, asi que "ParqueSur Fábrica" (con
    # tilde, como lo escribio quien respondio) normaliza a la misma clave
    # que "ParqueSur Fabrica" (el nombre canonico, sin tilde) -- sin este
    # alias explicito, _normaliza_centro no encuentra nada en el diccionario
    # y devuelve el texto original tal cual, dejando las dos variantes como
    # centros distintos en vez de fusionarlas.
    "parquesur fabrica": "ParqueSur Fabrica",
}

# Todas las tiendas/centros conocidos de Krispy Kreme — para poder registrar
# una salida manual incluso en un centro que todavía no tiene ninguna
# respuesta de Entrevista de Salida. Los nombres tienen que coincidir letra
# por letra con el campo de texto libre "Centro de Trabajo" del Excel de
# Entrevista de Salida — si no, una salida registrada aquí contaría como un
# centro distinto al que ya usan las respuestas reales. "ParqueSur - Leganés"
# ya no se usa: se dividió en "ParqueSur Fabrica" y "ParqueSur Tienda".
CENTROS_CONOCIDOS_KK = [
    "Caleido", "Gran Plaza 2", "La Gavia", "Oficina Central", "ParqueSur Fabrica", "ParqueSur Tienda", "Plenilunio",
    "Princesa",
]


def _normaliza_centro(centro):
    if not centro:
        return centro
    return _CENTRO_ALIASES.get(_normaliza_header(centro), centro)


def _normaliza_header(h):
    return (h or "").strip().lower().replace("í", "i").replace("ó", "o").replace("á", "a").replace(
        "é", "e"
    ).replace("ú", "u")


def _respuesta_con_valor(texto):
    limpio = (texto or "").strip()
    if not limpio:
        return False
    return limpio.lower() not in _RESPUESTAS_SIN_VALOR


# Igual que likertToNum() del script: "muy de acuerdo"/"muy en desacuerdo" se
# comprueban ANTES que "de acuerdo"/"en desacuerdo" a secas, porque los
# contienen como subcadena. "Totalmente de/en desacuerdo" y "Ni de acuerdo
# ni en desacuerdo" (las variantes que usa Clima Laboral) también se
# reconocen, por si algún formulario mezcla las dos redacciones.
def _likert_to_num(valor):
    if valor is None:
        return None
    s = _normaliza_header(str(valor))
    # El módulo de Test manda directamente el punto (1-5) en vez de la
    # leyenda — igual que _likert_points_strict en scoring_valores.py, esto
    # deja reconocer las dos formas sin duplicar el resto de la función.
    if s in ("1", "2", "3", "4", "5"):
        return int(s)
    if "muy en desacuerdo" in s or "totalmente en desacuerdo" in s:
        return 1
    if "muy de acuerdo" in s or "totalmente de acuerdo" in s:
        return 5
    if "en desacuerdo" in s:
        return 2
    if "ni de acuerdo" in s or "neutral" in s:
        return 3
    if "de acuerdo" in s:
        return 4
    return None


_BLOQUE_PREGUNTA_RE = re.compile(r"^(.*?):\s*\[(.+)\]\s*$")


def _bloque_label(prefijo):
    """"Indica tu nivel de acuerdo... sobre tu llegada a Saona" -> "Llegada
    a Saona" — se queda con lo que va después de "sobre" (sin el artículo
    inicial) para un título corto y legible, sin depender de qué empresa sea."""
    texto = re.sub(r"^.*?\bsobre\s+", "", prefijo, flags=re.IGNORECASE).strip()
    texto = re.sub(r"^(tu|tus|la|el|las|los)\s+", "", texto, flags=re.IGNORECASE)
    if not texto:
        return prefijo.strip()
    return texto[0].upper() + texto[1:]


# Las preguntas sueltas de escala sin bloque propio (ver _es_columna_likert)
# resultan ser siempre, en la práctica, sobre instalaciones/salario/horarios
# — es decir, "Condiciones Laborales", el mismo nombre de bloque que usan
# tanto Krispy Kreme como Saona en sus formularios.
GENERICO_BLOQUE_LABEL = "Condiciones Laborales"

# El texto de la categoría tal cual viene en el Excel a veces es una
# abreviatura de trabajo, no el nombre que se quiere mostrar (p.ej. el
# formulario dice "Experiencia Krispy Kreme" pero el nombre correcto es
# "Experiencia EN Krispy Kreme"). Normaliza solo esos casos conocidos; para
# cualquier categoría no listada aquí se usa el texto tal cual viene.
_BLOQUE_LABEL_ALIASES = {
    "experiencia krispy kreme": "Experiencia en Krispy Kreme",
    "cultura y ambiente en tu equipo": "Cultura y Ambiente",
    "condiciones laborales": "Condiciones Laborales",
}


def _normaliza_bloque_label(nombre):
    return _BLOQUE_LABEL_ALIASES.get(_normaliza_header(nombre), nombre)


def _tiene_categoria_en_texto(h):
    """True solo si el "." separa una categoría real de una pregunta real
    (formato Krispy Kreme: "Categoria.Pregunta") — una frase que simplemente
    termina en punto no cuenta: el partition dejaría la "pregunta" vacía."""
    if "." not in h:
        return False
    _, _, resto = h.partition(".")
    return len(resto.strip()) > 3


def _es_columna_likert(header, filas):
    """Para preguntas de escala SIN categoría en el encabezado ni formato
    "titulo: [pregunta]" (algunos formularios de KK traen preguntas sueltas,
    fuera de cualquier bloque tipo rejilla) — se decide mirando las
    respuestas reales: si la mayoría son valores Likert reconocibles, es una
    pregunta de escala (se agrupa en un bloque genérico), si no, es un
    comentario libre."""
    valores = [fila.get(header) for fila in filas if fila.get(header) not in (None, "")]
    if not valores:
        return False
    likert_vals = [v for v in valores if _likert_to_num(v) is not None]
    return len(likert_vals) / len(valores) >= 0.6


def _union_headers(filas):
    """Encabezados de TODAS las filas, no solo la primera — un Excel siempre
    tiene columnas homogéneas, pero filas que llegan del módulo de Test
    pueden traer preguntas distintas de una tanda a otra (p.ej. el enunciado
    se editó, o se mezclan respuestas de Excel con respuestas de Test). Si
    solo se mirara filas[0], cualquier columna que no estuviera en esa fila
    en concreto se perdería en silencio — ni se puntuaría ni aparecería como
    comentario."""
    vistos = set()
    resultado = []
    for fila in filas:
        for k in fila.keys():
            if k not in vistos:
                vistos.add(k)
                resultado.append(k)
    return resultado


def _mejor_columna_nombre(headers, metadata_cols):
    """La columna que _column_roles asigna a "nombre" es la PRIMERA que
    contenga esa palabra — en los Excel reales de Krispy Kreme eso suele ser
    una columna "Nombre" corta que llega vacía o con basura (un 1, resto de
    una celda de Forms), mientras que el nombre real y completo está en
    "Nombre y Apellido", una columna aparte que por eso mismo se clasificaba
    como comentario abierto en vez de como metadata. Para el cruce con
    Salidas Totales hace falta el nombre completo de verdad, así que aquí se
    prioriza cualquier encabezado que combine "nombre" y "apellido" antes de
    caer al comportamiento genérico de _column_roles."""
    for h in headers:
        norm = _normaliza_header(h)
        if "nombre" in norm and "apellido" in norm:
            return h
    return metadata_cols.get("nombre")


def _column_roles(headers, filas=None):
    """Dos formatos reconocidos para las preguntas de escala, además de
    preguntas sueltas sin bloque:
    - "titulo: [pregunta]" (Saona): el bloque es el texto antes de los dos
      puntos.
    - "Categoria.Pregunta" (Krispy Kreme, preguntas tipo rejilla): el bloque
      es el texto antes del punto.
    - Preguntas sueltas de escala sin ninguno de los dos formatos: se
      agrupan todas bajo un bloque genérico ("Otros aspectos"), detectadas
      por el contenido real de sus respuestas, no por texto hardcodeado.
    """
    filas = filas or []
    centro_col = None
    metadata_cols = {}
    motivo_col = None
    bloques = {}  # bloque_label (orden de aparición) -> [(header, pregunta), ...]
    abiertas = []

    for h in headers:
        norm = _normaliza_header(h)
        # "Primero que aparece gana" (no "el último"): con _union_headers
        # una oleada puede mezclar el encabezado real de un Excel con el
        # de una pregunta de Test redactada distinto (misma idea, otro
        # texto) — sin este orden, la columna nueva pisaría en silencio a
        # la que ya tenía datos reales acumulados.
        if ("centro de trabajo" in norm or "centro" in norm) and centro_col is None:
            centro_col = h
            continue
        if norm == "id":
            # Columna de fila autogenerada por Forms — no es ni metadata útil
            # ni comentario, se ignora del todo.
            continue
        if re.fullmatch(r"columna\d*", norm):
            # "Columna1", "Columna2"... es el nombre que pone Excel solo
            # cuando la celda de encabezado viene vacía — la pregunta real
            # perdió su texto al exportar. Aunque tenga datos de verdad
            # debajo, se ignora del todo (no entra en ningún bloque ni en
            # comentarios) hasta que se sepa qué pregunta era de verdad;
            # contarla con un rótulo sin sentido inflaría/desinflaría medias
            # de forma engañosa.
            continue
        matched_meta = False
        for hint, campo in METADATA_HINTS.items():
            if hint in norm and campo not in metadata_cols:
                metadata_cols[campo] = h
                matched_meta = True
                break
        if matched_meta:
            continue
        if "motivo" in norm and "salida" in norm and motivo_col is None:
            motivo_col = h
            continue
        m = _BLOQUE_PREGUNTA_RE.match(h)
        if m:
            prefijo, pregunta = m.group(1).strip(), m.group(2).strip()
            bloques.setdefault(_normaliza_bloque_label(_bloque_label(prefijo)), []).append((h, pregunta))
            continue
        if _tiene_categoria_en_texto(h):
            categoria, _, pregunta = h.partition(".")
            bloques.setdefault(_normaliza_bloque_label(categoria.strip()), []).append((h, pregunta.strip()))
            continue
        if _es_columna_likert(h, filas):
            bloques.setdefault(GENERICO_BLOQUE_LABEL, []).append((h, h.strip()))
            continue
        abiertas.append(h)

    return {
        "centro_col": centro_col,
        "metadata": metadata_cols,
        "motivo_col": motivo_col,
        "bloques": bloques,
        "abiertas": abiertas,
    }


# ======================= Cruce con "Salidas Totales" (cobertura/auditoría) =======================
# Puerto fiel del cruce fuzzy del script de Apps Script del usuario: agrupa
# "Salidas Totales" en secciones por centro (detectando el centro EN
# CUALQUIER COLUMNA de la fila título, porque el Excel real las trae
# desplazadas por celdas combinadas), y empareja cada respuesta con la
# salida real de esa persona por nombre (tokens con tolerancia a errores de
# tecleo), para poder asignar la fecha de baja real y calcular cobertura.

def _normaliza_nombre(s):
    s = _normaliza_header(str(s or ""))
    s = re.sub(r"[^a-z\s]", "", s)
    return re.sub(r"\s+", " ", s).strip()


def _levenshtein(a, b):
    m, n = len(a), len(b)
    dp = [[0] * (n + 1) for _ in range(m + 1)]
    for i in range(m + 1):
        dp[i][0] = i
    for j in range(n + 1):
        dp[0][j] = j
    for i in range(1, m + 1):
        for j in range(1, n + 1):
            costo = 0 if a[i - 1] == b[j - 1] else 1
            dp[i][j] = min(dp[i - 1][j] + 1, dp[i][j - 1] + 1, dp[i - 1][j - 1] + costo)
    return dp[m][n]


def _tokens_equivalentes(t1, t2):
    if t1 == t2:
        return True
    if len(t1) >= 3 and len(t2) >= 3 and (t2.startswith(t1) or t1.startswith(t2)):
        return True
    dist = _levenshtein(t1, t2)
    umbral = 1 if max(len(t1), len(t2)) <= 5 else 2
    return dist <= umbral


def _nombre_score(nombre_a, nombre_b):
    """1.0 = todos los tokens del nombre más corto encontraron pareja
    equivalente en el más largo (mismo criterio que exige el script: solo se
    acepta el cruce si el score es exactamente 1)."""
    tokens_a = _normaliza_nombre(nombre_a).split()
    tokens_b = _normaliza_nombre(nombre_b).split()
    if not tokens_a or not tokens_b:
        return 0
    corto, largo = (tokens_a, tokens_b) if len(tokens_a) <= len(tokens_b) else (tokens_b, tokens_a)
    usados = set()
    matched = 0
    for tc in corto:
        for i, tl in enumerate(largo):
            if i in usados:
                continue
            if _tokens_equivalentes(tc, tl):
                matched += 1
                usados.add(i)
                break
    return matched / len(corto)


def _parse_fecha(valor):
    if valor is None:
        return None
    if isinstance(valor, datetime):
        return valor
    if isinstance(valor, str):
        try:
            return datetime.fromisoformat(valor[:19])
        except ValueError:
            return None
    return None


def get_cuatrimestre(valor_fecha):
    """4 meses por tramo (1: ene-abr, 2: may-ago, 3: sep-dic) — igual que
    getCuatrimestre() del script. Etiqueta "nQaño" (ej. "2Q2026")."""
    dt = _parse_fecha(valor_fecha)
    if dt is None:
        return None
    n = 1 if dt.month <= 4 else (2 if dt.month <= 8 else 3)
    return {"label": f"{n}Q{dt.year}", "sort_key": f"{dt.year}-C{n}", "n": n, "year": dt.year}


def _sede_keywords_desde_centro(centro):
    """Palabras significativas (>=4 letras) del nombre del centro, para
    reconocer su sección dentro de "Salidas Totales" sin tener que
    hardcodear una lista de sedes por empresa — cualquier fila título que
    mencione alguna de estas palabras se asigna a ese centro."""
    tokens = re.findall(r"[a-z]+", _normaliza_header(centro))
    ignorar = {"saona", "krispy", "kreme", "centro", "trabajo", "tienda", "restaurante"}
    return [t for t in tokens if len(t) >= 4 and t not in ignorar]


# Puesto manda sobre la compañía: si el puesto es de producción/limpieza/
# decoración es Fábrica sin importar qué ponga la columna "Compañía" (hay
# una sola fábrica hoy, así que no hace falta más detalle) — así lo confirmó
# el usuario tras revisar los puestos reales de la hoja "Salidas Totales".
_PUESTOS_FABRICA = ("produccion", "limpieza", "decoracion")
# "Dependiente"/"Vendedor" (son lo mismo) y "Retail" (Jefe/a de Turno -
# Retail) son puestos de tienda — cuál tienda depende de la compañía.
_PUESTOS_TIENDA = ("dependiente", "vendedor", "retail")

_HORAS_SUFIJO_RE = re.compile(r"\s*\d+\s*h\s*$", re.IGNORECASE)

# Mismo patrón que boletines.py/boletines.js — validar aquí evita que un
# valor mal escrito en la columna de correo del Excel (típico: "N/A", un
# teléfono pegado por error) acabe colándose en el destinatario del
# "Enviar Recordatorio" (ver wireRecordatorio en entrevistas.js).
_EMAIL_RE = re.compile(r"^[^\s@]+@[^\s@]+\.[^\s@]+$")


def _puesto_canonico(puesto):
    """Unifica variantes del mismo puesto: quita el sufijo de horas del
    contrato ("Dependiente/a 20h" y "Dependiente/a 40h" son el mismo puesto,
    solo cambia la jornada) y renombra "Dependiente/a"/"Dependiente" a
    "Vendedor/a" (mismo rol, nombre más claro), tal como pidió el usuario."""
    if not puesto:
        return None
    texto = _HORAS_SUFIJO_RE.sub("", str(puesto).strip()).strip()
    texto = re.sub(r"\s+", " ", texto).replace("–", "-")
    if _normaliza_header(texto).startswith("dependiente"):
        return "Vendedor/a"
    return texto or None


def _tipo_centro(centro):
    """Clasifica un centro conocido en tienda/oficina/fábrica, para el
    resumen "Tienda X · Oficina Y · Fábrica Z"."""
    norm = _normaliza_header(centro or "")
    if "fabrica" in norm:
        return "fabrica"
    if "oficina" in norm:
        return "oficina"
    return "tienda"


def _centro_desde_compania_puesto(compania, puesto):
    """Traduce una fila de la hoja "Salidas Totales" en formato plano (con
    columnas Compañía + Puesto de trabajo, en vez del formato antiguo
    agrupado por secciones) a un centro conocido.
    - Puesto de producción/limpieza/decoración -> ParqueSur Fábrica, sin
      importar la compañía (hay una sola fábrica).
    - Puesto de tienda (dependiente/vendedor/retail) -> la tienda que
      indique la compañía ("T-MDxx COD-Nombre"), o ParqueSur Tienda por
      defecto si la compañía es la razón social genérica sin código.
    - Cualquier otro puesto bajo la razón social genérica (roles de
      oficina/corporativos, ej. "Programador", "Director Operaciones") ->
      Oficina Central.
    - Cualquier otro puesto bajo un código de tienda/fábrica concreto ->
      sin determinar (None), para no adivinar mal un caso no visto."""
    norm_puesto = _normaliza_header(puesto or "")
    norm_compania = _normaliza_header(compania or "")
    if any(k in norm_puesto for k in _PUESTOS_FABRICA):
        return "ParqueSur Fabrica"
    if any(k in norm_puesto for k in _PUESTOS_TIENDA):
        # "Glaseados Originales S.L." es la razón social genérica, sin
        # código de tienda concreto — para puestos de tienda bajo esa razón
        # social se asume ParqueSur Tienda por defecto (confirmado por el
        # usuario), a diferencia de "T-MDxx COD-Nombre" que sí trae la
        # tienda explícita.
        if "glaseados originales" in norm_compania or not compania:
            return "ParqueSur Tienda"
        nombre = compania.strip().rsplit("-", 1)[-1].strip()
        if not nombre:
            return "ParqueSur Tienda"
        if "parquesur" in _normaliza_header(nombre) or "parque sur" in _normaliza_header(nombre):
            return "ParqueSur Tienda"
        return nombre
    if "glaseados originales" in norm_compania:
        return "Oficina Central"
    return None


def _es_formato_plano_salidas(header):
    # "compa" y no "compania" a propósito: _normaliza_header no quita la ñ
    # (solo vocales acentuadas), así que "Compañía" normaliza a "compañía",
    # no a "compania" — "compa" es un prefijo seguro en cualquiera de las
    # dos formas.
    norm = [_normaliza_header(str(h)) for h in header if h]
    tiene = lambda palabra: any(palabra in h for h in norm)
    return tiene("nombre") and tiene("compa") and tiene("puesto") and tiene("fecha")


def _parse_salidas_totales_plano(ws):
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows_iter))
    except StopIteration:
        return []
    norm_header = [_normaliza_header(str(h)) if h else "" for h in header]
    idx_nombre = next(i for i, h in enumerate(norm_header) if "nombre" in h)
    idx_fecha = next(i for i, h in enumerate(norm_header) if "fecha" in h)
    idx_compania = next(i for i, h in enumerate(norm_header) if "compa" in h)
    idx_puesto = next(i for i, h in enumerate(norm_header) if "puesto" in h)
    idx_email = next((i for i, h in enumerate(norm_header) if "correo" in h or "email" in h), None)

    salidas = []
    for row in rows_iter:
        if row is None or all(c in (None, "") for c in row):
            continue
        nombre = row[idx_nombre] if idx_nombre < len(row) else None
        fecha_baja = row[idx_fecha] if idx_fecha < len(row) else None
        compania = row[idx_compania] if idx_compania < len(row) else None
        puesto = row[idx_puesto] if idx_puesto < len(row) else None
        email = row[idx_email] if idx_email is not None and idx_email < len(row) else None
        if not nombre or not hasattr(fecha_baja, "isoformat"):
            continue
        centro = _centro_desde_compania_puesto(compania, puesto)
        if not centro:
            continue
        email_limpio = str(email).strip() if email else None
        if email_limpio and not _EMAIL_RE.match(email_limpio):
            email_limpio = None
        salidas.append({
            "centro": centro, "nombre": str(nombre).strip(), "fecha_baja": fecha_baja.isoformat(),
            "puesto": _puesto_canonico(puesto),
            "email": email_limpio,
        })
    return salidas


def _parse_salidas_totales(ws, centros_conocidos):
    primera_fila = next(ws.iter_rows(values_only=True), None)
    if primera_fila and _es_formato_plano_salidas(primera_fila):
        return _parse_salidas_totales_plano(ws)

    sede_keywords = {c: _sede_keywords_desde_centro(c) for c in centros_conocidos}

    def _detectar_centro_en_fila(row):
        celdas = [str(c).strip() for c in row if c not in (None, "")]
        if not celdas:
            return None
        texto = _normaliza_header(" ".join(celdas))
        for centro, keywords in sede_keywords.items():
            if any(kw in texto for kw in keywords):
                return centro
        return None

    salidas = []
    centro_actual = None
    col_map = None
    for row in ws.iter_rows(values_only=True):
        row = list(row)
        no_vacias = sum(1 for c in row if c not in (None, ""))
        norm_cells = [_normaliza_header(str(c)) for c in row if c not in (None, "")]
        texto_fila = " ".join(norm_cells)
        has_trabajador = "trabajador" in texto_fila or "nombre" in texto_fila
        has_fecha_baja = "fecha" in texto_fila and "baja" in texto_fila
        centro_en_fila = _detectar_centro_en_fila(row)

        if no_vacias == 0:
            centro_actual = None
            col_map = None
            continue

        if centro_en_fila and not has_trabajador and no_vacias <= 2:
            centro_actual = centro_en_fila
            col_map = None
            continue

        if centro_actual and col_map is None and has_trabajador and has_fecha_baja:
            norm_por_col = [_normaliza_header(str(c)) if c not in (None, "") else "" for c in row]
            idx_nombre = next((i for i, c in enumerate(norm_por_col) if "trabajador" in c or "nombre" in c), None)
            idx_fecha = next((i for i, c in enumerate(norm_por_col) if "fecha" in c and "baja" in c), None)
            if idx_nombre is not None and idx_fecha is not None:
                col_map = {"nombre": idx_nombre, "fecha": idx_fecha}
            continue

        if centro_actual and col_map:
            nombre = row[col_map["nombre"]]
            fecha_baja = row[col_map["fecha"]]
            if nombre and hasattr(fecha_baja, "isoformat"):
                salidas.append({
                    "centro": centro_actual, "nombre": str(nombre).strip(), "fecha_baja": fecha_baja.isoformat(),
                    "puesto": None, "email": None,
                })

    return salidas


def ensure_entrevistas_tables():
    conn = get_connection()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS entrevistas_oleadas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            etiqueta TEXT,
            empresa TEXT NOT NULL DEFAULT 'kk',
            creado_en TEXT NOT NULL DEFAULT (datetime('now'))
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS entrevistas_respuestas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            oleada_id INTEGER NOT NULL REFERENCES entrevistas_oleadas(id),
            centro TEXT,
            fila_hash TEXT NOT NULL,
            datos_json TEXT NOT NULL,
            creado_en TEXT NOT NULL DEFAULT (datetime('now')),
            UNIQUE(oleada_id, fila_hash)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS entrevistas_importaciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            oleada_id INTEGER NOT NULL REFERENCES entrevistas_oleadas(id),
            archivo_nombre TEXT,
            subido_por TEXT,
            subido_en TEXT NOT NULL DEFAULT (datetime('now')),
            nuevas INTEGER NOT NULL DEFAULT 0,
            ya_existian INTEGER NOT NULL DEFAULT 0
        )
    """)
    cols_importaciones = {row[1] for row in conn.execute("PRAGMA table_info(entrevistas_importaciones)")}
    if "actualizadas" not in cols_importaciones:
        conn.execute("ALTER TABLE entrevistas_importaciones ADD COLUMN actualizadas INTEGER NOT NULL DEFAULT 0")
    # centro/nombre/fecha de baja/puesto/email — lo que hace falta para
    # calcular cobertura, cruzar con las respuestas, el resumen por puesto
    # de trabajo y poder mandar un recordatorio manual (botón "Enviar
    # Recordatorio", genera un mailto:, nunca se envía nada desde el
    # servidor). Nada de DNI ni otros datos sensibles de la hoja "Salidas
    # Totales" — el email se guarda porque hace falta para el recordatorio,
    # pero no se muestra en ninguna lista de nombres.
    conn.execute("""
        CREATE TABLE IF NOT EXISTS entrevistas_salidas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            oleada_id INTEGER NOT NULL REFERENCES entrevistas_oleadas(id),
            centro TEXT NOT NULL,
            nombre TEXT NOT NULL,
            fecha_baja TEXT NOT NULL
        )
    """)
    cols_salidas = {row[1] for row in conn.execute("PRAGMA table_info(entrevistas_salidas)")}
    if "puesto" not in cols_salidas:
        conn.execute("ALTER TABLE entrevistas_salidas ADD COLUMN puesto TEXT")
    if "email" not in cols_salidas:
        conn.execute("ALTER TABLE entrevistas_salidas ADD COLUMN email TEXT")
    # Override manual de un cruce respuesta<->salida — para cuando la misma
    # persona aparece en las dos auditorías (p.ej. "FLORES, LENIN MICHAEL" en
    # Salidas Totales vs "Lenin flores alvarado" en su propia respuesta) pero
    # el cruce automático por nombre no llega a 1.0 de score. Sin esto, la
    # única opción era "Registrar como salida" (que crea una salida NUEVA y
    # cuenta a la persona dos veces).
    conn.execute("""
        CREATE TABLE IF NOT EXISTS entrevistas_matches_manual (
            respuesta_id INTEGER PRIMARY KEY REFERENCES entrevistas_respuestas(id),
            salida_id INTEGER NOT NULL REFERENCES entrevistas_salidas(id)
        )
    """)
    conn.commit()
    conn.close()
    _migrar_alias_centro()


def _migrar_alias_centro():
    """Corrige datos que ya se importaron antes de existir _CENTRO_ALIASES —
    tanto la columna 'centro' como el propio valor embebido en datos_json
    (de ahí se vuelve a leer el centro cada vez que se calcula un reporte,
    así que hay que corregir los dos sitios para que no queden desalineados)."""
    conn = get_connection()
    filas = conn.execute("SELECT id, centro, datos_json FROM entrevistas_respuestas").fetchall()
    for fila in filas:
        nuevo_centro = _normaliza_centro(fila["centro"])
        datos = json.loads(fila["datos_json"])
        cambiado = nuevo_centro != fila["centro"]
        for k, v in datos.items():
            if isinstance(v, str):
                v_normalizado = _normaliza_centro(v)
                if v_normalizado != v:
                    datos[k] = v_normalizado
                    cambiado = True
        if cambiado:
            conn.execute(
                "UPDATE entrevistas_respuestas SET centro = ?, datos_json = ? WHERE id = ?",
                (nuevo_centro, json.dumps(datos, ensure_ascii=False, default=str), fila["id"]),
            )
    # Antes esto se hacia con "WHERE LOWER(centro) = ?" contra la clave del
    # alias -- LOWER() de SQLite no quita acentos, asi que una variante como
    # "ParqueSur Fábrica" (con tilde) nunca coincidia con la clave
    # "parquesur fabrica" (sin tilde) y se quedaba sin migrar. Se usa la
    # misma normalizacion en Python que ya aplica arriba a las respuestas.
    salidas = conn.execute("SELECT id, centro FROM entrevistas_salidas").fetchall()
    for s in salidas:
        nuevo_centro = _normaliza_centro(s["centro"])
        if nuevo_centro != s["centro"]:
            conn.execute("UPDATE entrevistas_salidas SET centro = ? WHERE id = ?", (nuevo_centro, s["id"]))
    conn.commit()
    conn.close()


def list_centros_conocidos(empresa="kk"):
    """Todas las tiendas/centros conocidos de la empresa, no solo los que ya
    tienen alguna respuesta — para poder dar de alta una salida manual en un
    centro que todavía no aparece en ninguna oleada."""
    if empresa == "kk":
        return CENTROS_CONOCIDOS_KK
    # Saona no tiene todavía un listado maestro de restaurantes (a diferencia
    # de KK) — se usa la unión de todos los centros vistos alguna vez, tanto
    # en respuestas como en salidas, de cualquier oleada de la empresa.
    conn = get_connection()
    rows = conn.execute("""
        SELECT DISTINCT centro FROM (
            SELECT r.centro AS centro FROM entrevistas_respuestas r
            JOIN entrevistas_oleadas o ON o.id = r.oleada_id
            WHERE o.empresa = ? AND r.centro IS NOT NULL
            UNION
            SELECT s.centro AS centro FROM entrevistas_salidas s
            JOIN entrevistas_oleadas o ON o.id = s.oleada_id
            WHERE o.empresa = ?
        ) ORDER BY centro
    """, (empresa, empresa)).fetchall()
    conn.close()
    return [r["centro"] for r in rows]


def _hash_fila(fila):
    blob = json.dumps(fila, sort_keys=True, ensure_ascii=False, default=str)
    return hashlib.sha256(blob.encode("utf-8")).hexdigest()


def _clave_identidad(fila, nombre_col, hora_col):
    nombre = fila.get(nombre_col)
    hora = fila.get(hora_col)
    if not nombre or not hora:
        return None
    return (_normaliza_header(str(nombre)).strip(), str(hora).strip())


def _read_sheet_rows(ws):
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        return []
    filas = []
    for row in rows_iter:
        if row is None or all(v is None for v in row):
            continue
        datos = {}
        for i, key in enumerate(header):
            if not key:
                continue
            value = row[i] if i < len(row) else None
            if hasattr(value, "isoformat"):
                value = value.isoformat()
            datos[key] = value
        if datos:
            filas.append(datos)
    return filas


def get_or_create_oleada(nueva, etiqueta=None, empresa="kk"):
    conn = get_connection()
    if not nueva:
        row = conn.execute(
            "SELECT id FROM entrevistas_oleadas WHERE empresa = ? ORDER BY id DESC LIMIT 1", (empresa,)
        ).fetchone()
        if row:
            oleada_id = row[0]
            conn.close()
            return oleada_id
    cur = conn.execute("INSERT INTO entrevistas_oleadas (etiqueta, empresa) VALUES (?, ?)", (etiqueta, empresa))
    conn.commit()
    oleada_id = cur.lastrowid
    conn.close()
    return oleada_id


def get_oleada_empresa(oleada_id):
    conn = get_connection()
    row = conn.execute("SELECT empresa FROM entrevistas_oleadas WHERE id = ?", (oleada_id,)).fetchone()
    conn.close()
    return row["empresa"] if row else None


def import_excel(file_bytes, archivo_nombre, subido_por, nueva_oleada=False, empresa="kk"):
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        raise ValueError(f"No se pudo leer el archivo Excel: {e}")

    if "Respuestas" not in wb.sheetnames:
        raise ValueError("El Excel debe tener una hoja llamada 'Respuestas'")

    filas = _read_sheet_rows(wb["Respuestas"])
    if not filas:
        raise ValueError("La hoja 'Respuestas' no tiene filas de datos")

    oleada_id = get_or_create_oleada(nueva_oleada, empresa=empresa)

    conn = get_connection()
    cur = conn.execute(
        "INSERT INTO entrevistas_importaciones (oleada_id, archivo_nombre, subido_por) VALUES (?, ?, ?)",
        (oleada_id, archivo_nombre, subido_por),
    )
    importacion_id = cur.lastrowid

    roles = _column_roles(_union_headers(filas), filas)
    centro_col = roles["centro_col"]
    nombre_col = _mejor_columna_nombre(_union_headers(filas), roles["metadata"])
    hora_col = roles["metadata"].get("hora_inicio")

    # La identidad de una respuesta es la persona + el momento exacto en que
    # la envió (nombre + hora de inicio), NO la fila completa. Si el Excel de
    # origen se corrige después (p. ej. al dividir "ParqueSur - Leganés" en
    # Fábrica/Tienda, cambiando la columna de centro de gente que ya había
    # respondido), una reimportación debe ACTUALIZAR esa respuesta con el
    # dato corregido, no duplicarla — antes, como el hash se calculaba sobre
    # la fila entera, cualquier corrección hacía que la fila pareciera
    # "nueva" y se insertaba una segunda copia de la misma persona.
    existentes_por_identidad = {}
    if nombre_col and hora_col:
        for id_existente, datos_existentes in _fetch_respuestas(conn, oleada_id, None):
            clave = _clave_identidad(datos_existentes, nombre_col, hora_col)
            if clave:
                existentes_por_identidad[clave] = id_existente

    nuevas = 0
    actualizadas = 0
    ya_existian = 0
    centros_vistos = set()
    for fila in filas:
        centro = _normaliza_centro(str(fila.get(centro_col)).strip()) if centro_col and fila.get(centro_col) else None
        if centro_col:
            fila[centro_col] = centro
        if centro:
            centros_vistos.add(centro)
        fila_hash = _hash_fila(fila)
        datos_json = json.dumps(fila, ensure_ascii=False, default=str)

        clave = _clave_identidad(fila, nombre_col, hora_col) if nombre_col and hora_col else None
        if clave and clave in existentes_por_identidad:
            id_existente = existentes_por_identidad[clave]
            conn.execute(
                "UPDATE entrevistas_respuestas SET centro = ?, fila_hash = ?, datos_json = ? WHERE id = ?",
                (centro, fila_hash, datos_json, id_existente),
            )
            actualizadas += 1
            continue

        # Sin nombre+hora que identifiquen la fila (formulario sin esas
        # columnas), cae al criterio anterior: hash exacto de toda la fila.
        existe = conn.execute(
            "SELECT id FROM entrevistas_respuestas WHERE oleada_id = ? AND fila_hash = ?", (oleada_id, fila_hash)
        ).fetchone()
        if existe:
            ya_existian += 1
            continue
        cur_fila = conn.execute(
            "INSERT INTO entrevistas_respuestas (oleada_id, centro, fila_hash, datos_json) VALUES (?, ?, ?, ?)",
            (oleada_id, centro, fila_hash, datos_json),
        )
        nuevas += 1
        if clave:
            existentes_por_identidad[clave] = cur_fila.lastrowid

    # "Salidas Totales" (si viene en el mismo Excel) es una foto completa de
    # las bajas reales — se reemplaza entera en cada importación, no se
    # acumula, para que siempre refleje la lista más reciente.
    salidas_nuevas = 0
    if "Salidas Totales" in wb.sheetnames and centros_vistos:
        salidas = _parse_salidas_totales(wb["Salidas Totales"], centros_vistos)
        conn.execute("DELETE FROM entrevistas_salidas WHERE oleada_id = ?", (oleada_id,))
        for s in salidas:
            conn.execute(
                "INSERT INTO entrevistas_salidas (oleada_id, centro, nombre, fecha_baja, puesto, email) VALUES (?, ?, ?, ?, ?, ?)",
                (oleada_id, s["centro"], s["nombre"], s["fecha_baja"], s.get("puesto"), s.get("email")),
            )
        salidas_nuevas = len(salidas)

    conn.execute(
        "UPDATE entrevistas_importaciones SET nuevas = ?, ya_existian = ?, actualizadas = ? WHERE id = ?",
        (nuevas, ya_existian, actualizadas, importacion_id),
    )
    conn.commit()
    conn.close()
    return {
        "oleada_id": oleada_id, "nuevas": nuevas, "ya_existian": ya_existian, "actualizadas": actualizadas,
        "total_en_excel": len(filas), "salidas_totales": salidas_nuevas,
    }


def ingest_fila_directa(empresa, fila, origen="Formulario web"):
    """Alimenta Entrevista de Salida con UNA fila recién recibida (desde el
    módulo de Test cuando alguien envía el formulario público) exactamente
    igual que si viniera de un Excel: mismo detector de bloques/preguntas
    por el texto del encabezado, mismo dedup por hash — así el módulo de
    Test puede sustituir por completo al Excel de Forms para este formulario
    también. IMPORTANTE: para que una pregunta de escala caiga en el bloque
    correcto, su enunciado en el constructor de Test debe seguir el mismo
    formato que ya reconoce _column_roles: "<Bloque>: [<Pregunta>]" (formato
    Saona) o "<Categoria>.<Pregunta>" (formato Krispy Kreme) — un enunciado
    libre sin ese formato cae en el bloque genérico o se trata como
    comentario abierto, igual que pasaría con un Excel mal formado."""
    oleada_id = get_or_create_oleada(nueva=False, empresa=empresa)

    conn = get_connection()
    cur = conn.execute(
        "INSERT INTO entrevistas_importaciones (oleada_id, archivo_nombre, subido_por, nuevas, ya_existian) "
        "VALUES (?, ?, ?, 0, 0)",
        (oleada_id, origen, "sistema"),
    )
    importacion_id = cur.lastrowid

    roles = _column_roles(list(fila.keys()), [fila])
    centro_col = roles["centro_col"]
    centro = _normaliza_centro(str(fila.get(centro_col)).strip()) if centro_col and fila.get(centro_col) else None
    if centro_col:
        fila[centro_col] = centro

    fila_hash = _hash_fila(fila)
    existe = conn.execute(
        "SELECT id FROM entrevistas_respuestas WHERE oleada_id = ? AND fila_hash = ?", (oleada_id, fila_hash)
    ).fetchone()
    nuevas = 0
    ya_existian = 0
    if existe:
        ya_existian = 1
    else:
        datos_json = json.dumps(fila, ensure_ascii=False, default=str)
        conn.execute(
            "INSERT INTO entrevistas_respuestas (oleada_id, centro, fila_hash, datos_json) VALUES (?, ?, ?, ?)",
            (oleada_id, centro, fila_hash, datos_json),
        )
        nuevas = 1

    conn.execute(
        "UPDATE entrevistas_importaciones SET nuevas = ?, ya_existian = ? WHERE id = ?",
        (nuevas, ya_existian, importacion_id),
    )
    conn.commit()
    conn.close()
    return nuevas


def list_oleadas(empresa="kk"):
    conn = get_connection()
    rows = conn.execute("""
        SELECT o.id, o.etiqueta, o.creado_en, COUNT(r.id) AS num_respuestas,
               ROW_NUMBER() OVER (ORDER BY o.id ASC) AS numero
        FROM entrevistas_oleadas o
        LEFT JOIN entrevistas_respuestas r ON r.oleada_id = o.id
        WHERE o.empresa = ?
        GROUP BY o.id ORDER BY o.id DESC
    """, (empresa,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def list_centros(oleada_id):
    conn = get_connection()
    rows = conn.execute(
        "SELECT DISTINCT centro FROM entrevistas_respuestas WHERE oleada_id = ? AND centro IS NOT NULL ORDER BY centro",
        (oleada_id,),
    ).fetchall()
    conn.close()
    return [r["centro"] for r in rows]


def _fetch_respuestas(conn, oleada_id, centro):
    """Devuelve [(respuesta_id, datos), ...] — el id hace falta para poder
    guardar un match manual contra una respuesta concreta en compute_evolucion;
    compute_reporte simplemente lo ignora."""
    if centro:
        rows = conn.execute(
            "SELECT id, datos_json FROM entrevistas_respuestas WHERE oleada_id = ? AND centro = ?", (oleada_id, centro)
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT id, datos_json FROM entrevistas_respuestas WHERE oleada_id = ?", (oleada_id,)
        ).fetchall()
    return [(r["id"], json.loads(r["datos_json"])) for r in rows]


def list_respuestas_con_motivo(oleada_id, centro=None):
    """Una fila por respuesta (nombre + motivo de salida tal cual esta
    registrado hoy) -- para poder corregir a mano el motivo cuando el que
    marco la persona en el formulario no refleja el motivo real de su
    salida (ver update_motivo)."""
    conn = get_connection()
    filas_con_id = _fetch_respuestas(conn, oleada_id, centro)
    conn.close()
    if not filas_con_id:
        return []
    todas_las_filas = [datos for _id, datos in filas_con_id]
    roles = _column_roles(_union_headers(todas_las_filas), todas_las_filas)
    motivo_col = roles["motivo_col"]
    nombre_col = _mejor_columna_nombre(_union_headers(todas_las_filas), roles["metadata"])
    centro_col = roles["centro_col"]
    return [
        {
            "id": respuesta_id,
            "nombre": (fila.get(nombre_col) or "").strip() if nombre_col else None,
            "centro": fila.get(centro_col) if centro_col else None,
            "motivo": fila.get(motivo_col) if motivo_col else None,
        }
        for respuesta_id, fila in filas_con_id
    ]


def update_motivo(respuesta_id, nuevo_motivo):
    """Corrige el motivo de salida de UNA respuesta concreta -- el motivo que
    la persona marco en el formulario no siempre coincide con el motivo real
    de su baja, y esto permite a RRHH ajustarlo sin tener que reimportar el
    Excel entero. Escribe directamente en datos_json (misma columna que ya
    detecta _column_roles como "motivo"), asi compute_reporte lo recoge en el
    siguiente calculo sin cambios adicionales."""
    conn = get_connection()
    row = conn.execute(
        "SELECT oleada_id, datos_json FROM entrevistas_respuestas WHERE id = ?", (respuesta_id,)
    ).fetchone()
    if row is None:
        conn.close()
        raise ValueError("Respuesta no encontrada")

    todas_las_filas = [datos for _id, datos in _fetch_respuestas(conn, row["oleada_id"], None)]
    roles = _column_roles(_union_headers(todas_las_filas), todas_las_filas)
    motivo_col = roles["motivo_col"]
    if not motivo_col:
        conn.close()
        raise ValueError("Esta oleada no tiene una columna de motivo de salida reconocida")

    datos = json.loads(row["datos_json"])
    datos[motivo_col] = nuevo_motivo
    conn.execute(
        "UPDATE entrevistas_respuestas SET datos_json = ? WHERE id = ?",
        (json.dumps(datos, ensure_ascii=False, default=str), respuesta_id),
    )
    conn.commit()
    conn.close()


def compute_reporte(oleada_id, centro=None):
    conn = get_connection()
    filas_con_id = _fetch_respuestas(conn, oleada_id, centro)
    conn.close()
    if not filas_con_id:
        raise ValueError("No hay respuestas para este centro en esta oleada")
    filas = [datos for _id, datos in filas_con_id]

    roles = _column_roles(_union_headers(filas), filas)
    n = len(filas)

    suma_global = 0
    cuenta_global = 0
    bloques = []
    for bloque_nombre, preguntas in roles["bloques"].items():
        preguntas_detalle = []
        suma_bloque = 0.0
        count_bloque = 0
        for header, texto_pregunta in preguntas:
            valores = [v for v in (_likert_to_num(fila.get(header)) for fila in filas) if v is not None]
            if not valores:
                continue
            promedio_p = sum(valores) / len(valores)
            preguntas_detalle.append({"pregunta": texto_pregunta, "promedio": round(promedio_p, 2)})
            suma_bloque += promedio_p
            count_bloque += 1
            suma_global += sum(valores)
            cuenta_global += len(valores)
        promedio_bloque = round(suma_bloque / count_bloque, 2) if count_bloque else None
        bloques.append({"nombre": bloque_nombre, "promedio": promedio_bloque, "preguntas": preguntas_detalle})

    satisfaccion_general = round(suma_global / cuenta_global, 2) if cuenta_global else None

    motivos_map = {}
    if roles["motivo_col"]:
        for fila in filas:
            m = fila.get(roles["motivo_col"]) or "Sin dato"
            m = str(m).strip() or "Sin dato"
            motivos_map[m] = motivos_map.get(m, 0) + 1
    motivos = sorted(
        [{"motivo": m, "cantidad": c, "porcentaje": round(c / n * 100, 1)} for m, c in motivos_map.items()],
        key=lambda x: -x["cantidad"],
    )

    abiertas = {}
    for header in roles["abiertas"]:
        # "Puesto de trabajo Tienda/Oficina/Fábrica/..." son preguntas
        # condicionales del propio test (solo se le muestra al respondiente
        # la que corresponde a su centro) — no son un comentario libre, y
        # esa info ya se resume en la sección "Puestos de trabajo" (aunque
        # viene de una fuente distinta, Salidas Totales). Se excluyen de
        # Comentarios abiertos para no duplicar/fragmentar la vista.
        if _normaliza_header(header).startswith("puesto de trabajo"):
            continue
        textos = [str(fila[header]).strip() for fila in filas if fila.get(header)]
        textos = [t for t in textos if _respuesta_con_valor(t)]
        # Preguntas que se eliminaron del formulario antes de que nadie llegara
        # a contestarlas no aportan nada (siempre salen "sin comentarios") —
        # se omiten en vez de mostrar una tarjeta vacía.
        if textos:
            abiertas[header] = textos

    # N (quién respondió) vs Total (quién debía responder, según Salidas
    # Totales) son números distintos — antes solo se mostraba N y se leía
    # como si fuera el total. tiene_salidas_totales indica si hay datos de
    # Salidas Totales para poder calcular Total/Participación (si nunca se
    # subió esa hoja, se queda en None en vez de mostrar un 0 engañoso).
    conn2 = get_connection()
    salidas = _fetch_salidas(conn2, oleada_id, centro)
    conn2.close()
    total_salidas = len(salidas)
    tiene_salidas_totales = total_salidas > 0
    participacion = round(n / total_salidas * 100, 1) if total_salidas else None

    salidas_por_tipo = {"tienda": 0, "oficina": 0, "fabrica": 0}
    puestos_map = {"tienda": {}, "oficina": {}, "fabrica": {}}
    for s in salidas:
        tipo = _tipo_centro(s["centro"])
        salidas_por_tipo[tipo] += 1
        if s["puesto"]:
            puestos_map[tipo][s["puesto"]] = puestos_map[tipo].get(s["puesto"], 0) + 1
    puestos_por_tipo = {
        tipo: sorted(
            [{"puesto": p, "cantidad": c} for p, c in mapa.items()],
            key=lambda x: -x["cantidad"],
        )
        for tipo, mapa in puestos_map.items()
    }

    return {
        "oleada_id": oleada_id,
        "centro": centro,
        "n": n,
        "total_salidas": total_salidas if tiene_salidas_totales else None,
        "tiene_salidas_totales": tiene_salidas_totales,
        "participacion": participacion,
        "salidas_por_tipo": salidas_por_tipo if tiene_salidas_totales else None,
        "puestos_por_tipo": puestos_por_tipo,
        "satisfaccion_general": satisfaccion_general,
        "bloques": bloques,
        "motivos": motivos,
        "abiertas": abiertas,
    }


def _fetch_salidas(conn, oleada_id, centro):
    if centro:
        rows = conn.execute(
            "SELECT id, centro, nombre, fecha_baja, puesto, email FROM entrevistas_salidas WHERE oleada_id = ? AND centro = ?",
            (oleada_id, centro),
        ).fetchall()
    else:
        rows = conn.execute(
            "SELECT id, centro, nombre, fecha_baja, puesto, email FROM entrevistas_salidas WHERE oleada_id = ?", (oleada_id,)
        ).fetchall()
    return [dict(r) for r in rows]


def _fetch_matches_manual(conn, respuesta_ids):
    if not respuesta_ids:
        return {}
    placeholders = ",".join("?" for _ in respuesta_ids)
    rows = conn.execute(
        f"SELECT respuesta_id, salida_id FROM entrevistas_matches_manual WHERE respuesta_id IN ({placeholders})",
        respuesta_ids,
    ).fetchall()
    return {r["respuesta_id"]: r["salida_id"] for r in rows}


def set_match_manual(respuesta_id, salida_id):
    """Fuerza que esta respuesta concreta cuente como la salida indicada —
    para el caso de dos entradas que en realidad son la misma persona pero
    el cruce automático por nombre no llegó a 1.0 de score."""
    conn = get_connection()
    conn.execute(
        "INSERT INTO entrevistas_matches_manual (respuesta_id, salida_id) VALUES (?, ?) "
        "ON CONFLICT(respuesta_id) DO UPDATE SET salida_id = excluded.salida_id",
        (respuesta_id, salida_id),
    )
    conn.commit()
    conn.close()


def quitar_match_manual(respuesta_id):
    conn = get_connection()
    conn.execute("DELETE FROM entrevistas_matches_manual WHERE respuesta_id = ?", (respuesta_id,))
    conn.commit()
    conn.close()


def borrar_respuesta(respuesta_id):
    """Elimina una respuesta de Entrevista de Salida por completo (p.ej. un
    registro de prueba) — junto con su vínculo manual a una salida si tenía
    uno, para no dejar una fila huérfana en entrevistas_matches_manual."""
    conn = get_connection()
    conn.execute("DELETE FROM entrevistas_matches_manual WHERE respuesta_id = ?", (respuesta_id,))
    conn.execute("DELETE FROM entrevistas_respuestas WHERE id = ?", (respuesta_id,))
    conn.commit()
    conn.close()


def add_salida(oleada_id, centro, nombre, fecha_baja):
    """Registra una salida real a mano, sin depender del Excel de "Salidas
    Totales" — sirve tanto para dar de alta a alguien que no vino en la hoja
    como para confirmar una respuesta que no cruzó automáticamente (se le
    pasa su propio nombre/centro/fecha y queda contada igual que si viniera
    del Excel)."""
    conn = get_connection()
    conn.execute(
        "INSERT INTO entrevistas_salidas (oleada_id, centro, nombre, fecha_baja) VALUES (?, ?, ?, ?)",
        (oleada_id, centro, nombre, fecha_baja),
    )
    conn.commit()
    conn.close()


def compute_evolucion(oleada_id, centro=None):
    """Evolución por cuatrimestre — equivalente a Resumen_General del
    script: cruza cada respuesta con su baja real (por centro + nombre) para
    saber a qué cuatrimestre pertenece de verdad, y a partir de ahí calcula
    Total Ponderado / Ítem Mínimo-Máximo por bloque y cobertura, cuatrimestre
    a cuatrimestre."""
    conn = get_connection()
    filas_con_id = _fetch_respuestas(conn, oleada_id, centro)
    salidas = _fetch_salidas(conn, oleada_id, centro)
    matches_manual = _fetch_matches_manual(conn, [rid for rid, _ in filas_con_id])
    conn.close()
    if not filas_con_id:
        raise ValueError("No hay respuestas para este centro en esta oleada")

    _todas_las_filas = [datos for _id, datos in filas_con_id]
    roles = _column_roles(_union_headers(_todas_las_filas), _todas_las_filas)
    centro_col = roles["centro_col"]
    nombre_col = _mejor_columna_nombre(_union_headers(_todas_las_filas), roles["metadata"])
    fecha_col = roles["metadata"].get("hora_inicio")

    salidas_por_id = {s["id"]: s for s in salidas}
    salidas_por_centro = {}
    for s in salidas:
        salidas_por_centro.setdefault(s["centro"], []).append(s)

    usadas_ids = set()  # ids de entrevistas_salidas ya usados (fuzzy o manual)
    enriquecidas = []
    auditoria_g = []
    for respuesta_id, fila in filas_con_id:
        centro_fila = fila.get(centro_col) if centro_col else None
        nombre_fila = fila.get(nombre_col) if nombre_col else None
        fecha_propia = fila.get(fecha_col) if fecha_col else None
        fecha_efectiva = fecha_propia
        matched = False

        # Un match manual (ver set_match_manual) tiene prioridad absoluta
        # sobre el cruce automático por nombre — es la corrección explícita
        # de un humano para un caso donde el fuzzy-match no llegó a 1.0.
        salida_manual_id = matches_manual.get(respuesta_id)
        if salida_manual_id is not None and salida_manual_id in salidas_por_id and salida_manual_id not in usadas_ids:
            s = salidas_por_id[salida_manual_id]
            usadas_ids.add(salida_manual_id)
            fecha_efectiva = s["fecha_baja"]
            matched = True
        else:
            for s in salidas_por_centro.get(centro_fila, []):
                if s["id"] in usadas_ids:
                    continue
                if nombre_fila and _nombre_score(nombre_fila, s["nombre"]) == 1:
                    usadas_ids.add(s["id"])
                    fecha_efectiva = s["fecha_baja"]
                    matched = True
                    break

        if not matched:
            auditoria_g.append({"respuesta_id": respuesta_id, "nombre": nombre_fila, "fecha": fecha_propia, "centro": centro_fila})

        cuat = get_cuatrimestre(fecha_efectiva)
        enriquecidas.append({"fila": fila, "cuatrimestre": cuat, "matched": matched})

    auditoria_f = [
        {"salida_id": s["id"], "nombre": s["nombre"], "fecha_baja": s["fecha_baja"], "centro": s["centro"], "email": s["email"]}
        for s in salidas
        if s["id"] not in usadas_ids
    ]

    periodos_map = {}
    for e in enriquecidas:
        if not e["cuatrimestre"]:
            continue
        sk = e["cuatrimestre"]["sort_key"]
        periodos_map.setdefault(sk, {"label": e["cuatrimestre"]["label"], "sort_key": sk, "filas": []})
        periodos_map[sk]["filas"].append(e["fila"])
    periodos = sorted(periodos_map.values(), key=lambda p: p["sort_key"])

    stats_por_periodo = []
    for p in periodos:
        bloques_stats = []
        for bloque_nombre, preguntas in roles["bloques"].items():
            items = []
            for header, texto_pregunta in preguntas:
                valores = [v for v in (_likert_to_num(f.get(header)) for f in p["filas"]) if v is not None]
                media = sum(valores) / len(valores) if valores else None
                items.append({"pregunta": texto_pregunta, "media": media, "n": len(valores)})
            validos = [it for it in items if it["media"] is not None]
            total_n = sum(it["n"] for it in validos)
            total_ponderado = (sum(it["media"] * it["n"] for it in validos) / total_n) if total_n else None
            bloques_stats.append({
                "nombre": bloque_nombre,
                "total_ponderado": round(total_ponderado, 2) if total_ponderado is not None else None,
                "items": items,
            })
        min_max = []
        for bl in bloques_stats:
            validos = [it for it in bl["items"] if it["media"] is not None]
            if not validos:
                min_max.append({"bloque": bl["nombre"], "item_min": None, "min": None, "item_max": None, "max": None})
                continue
            item_min = min(validos, key=lambda it: it["media"])
            item_max = max(validos, key=lambda it: it["media"])
            min_max.append({
                "bloque": bl["nombre"],
                "item_min": item_min["pregunta"], "min": round(item_min["media"], 2),
                "item_max": item_max["pregunta"], "max": round(item_max["media"], 2),
            })
        stats_por_periodo.append({
            "label": p["label"], "sort_key": p["sort_key"], "n": len(p["filas"]),
            "bloques": bloques_stats, "min_max": min_max,
        })

    cobertura_map = {}
    for s in salidas:
        cuat = get_cuatrimestre(s["fecha_baja"])
        if not cuat:
            continue
        cobertura_map.setdefault(cuat["sort_key"], {"label": cuat["label"], "sort_key": cuat["sort_key"], "debieron": 0, "respondieron": 0})
        cobertura_map[cuat["sort_key"]]["debieron"] += 1
    for e in enriquecidas:
        if not e["matched"] or not e["cuatrimestre"]:
            continue
        sk = e["cuatrimestre"]["sort_key"]
        cobertura_map.setdefault(sk, {"label": e["cuatrimestre"]["label"], "sort_key": sk, "debieron": 0, "respondieron": 0})
        cobertura_map[sk]["respondieron"] += 1
    cobertura = sorted(cobertura_map.values(), key=lambda c: c["sort_key"])
    for c in cobertura:
        c["porcentaje"] = round(c["respondieron"] / c["debieron"] * 100, 1) if c["debieron"] else None

    return {
        "centro": centro,
        "tiene_salidas_totales": len(salidas) > 0,
        "total_salidas": len(salidas),
        "stats_por_periodo": stats_por_periodo,
        "cobertura": cobertura,
        "auditoria_f": auditoria_f,
        "auditoria_g": auditoria_g,
    }


ensure_entrevistas_tables()
