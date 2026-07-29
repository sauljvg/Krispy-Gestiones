import hashlib
import io
import json
import re

from openpyxl import load_workbook

from db import get_connection

LIKERT_ORDEN = ["Totalmente de acuerdo", "De acuerdo", "Neutral", "En desacuerdo", "Totalmente en desacuerdo"]

# Algunos formularios (Saona, Google Forms) usan "Ni de acuerdo ni en
# desacuerdo" en vez de "Neutral" para el mismo nivel intermedio — sin este
# alias esas respuestas no encajaban en ningún cubo de LIKERT_ORDEN y se
# perdían silenciosamente del conteo.
LIKERT_ALIASES = {
    "ni de acuerdo ni en desacuerdo": "Neutral",
}


def _canonicalizar_likert(valor):
    if valor in LIKERT_ORDEN:
        return valor
    return LIKERT_ALIASES.get(str(valor or "").strip().lower())


METADATA_HINTS = {
    "id": "id",
    "hora de inicio": "hora_inicio",
    "hora de finalizacion": "hora_fin",
    "hora de finalización": "hora_fin",
    "marca temporal": "hora_inicio",
    "correo electronico": "correo",
    "correo electrónico": "correo",
    "nombre": "nombre",
}

STOPWORDS = {
    # artículos, preposiciones, conjunciones
    "de", "la", "el", "en", "que", "y", "a", "los", "las", "un", "una", "mi", "me", "se", "por", "con",
    "es", "lo", "del", "al", "su", "sus", "para", "mas", "más", "esta", "este", "estas", "estos", "esa",
    "ese", "esos", "esas", "eso", "esto", "aquello", "aquel", "aquella", "aquellos", "aquellas",
    "hay", "no", "si", "sí", "o", "unos", "unas", "le", "les", "nos", "os", "ya", "todo", "toda", "todos",
    "todas", "como", "pero", "tambien", "también", "porque", "cuando", "donde", "sobre", "entre", "sin",
    "desde", "hasta", "hacia", "durante", "mediante", "según", "segun", "sino", "pues", "aunque",
    "mientras", "e", "u", "ni", "tan", "tanto", "tanta", "tantos", "tantas",
    # pronombres
    "yo", "tu", "tú", "él", "ella", "ellos", "ellas", "nosotros", "nosotras", "vosotros", "vosotras",
    "usted", "ustedes", "cual", "cuales", "quien", "quienes", "cuyo", "cuya", "cuyos", "cuyas",
    "mismo", "misma", "mismos", "mismas", "otro", "otra", "otros", "otras", "algo", "alguien",
    "alguna", "algunas", "alguno", "algunos", "nada", "nadie", "algún", "ningún", "ninguno", "ninguna",
    # verbos comunes (ser/estar/haber/tener/poder/deber/hacer/querer) en formas frecuentes
    "ser", "estar", "hacer", "tener", "poder", "deber", "querer", "creo", "considero", "creemos",
    "sea", "son", "soy", "eres", "somos", "sois", "fue", "fueron", "era", "eran", "siendo", "sido",
    "esta", "estan", "están", "estamos", "estoy", "estaba", "estaban", "estuvo",
    "hace", "hacen", "hacemos", "hago", "haciendo", "hecho",
    "tiene", "tienen", "tenemos", "tengo", "tenía", "tenia",
    "puede", "pueden", "podemos", "puedo", "podría", "podria", "podrían", "podrian",
    "debe", "deben", "debemos", "debería", "deberia", "deberían", "deberian",
    "quiere", "quieren", "queremos", "quiero", "quisiera",
    "seguir", "sigue", "siguen", "seguimos", "sigo",
    # adverbios/cuantificadores genéricos poco informativos
    "poco", "pocos", "poca", "pocas", "mucho", "muchos", "mucha", "muchas", "cada", "vez", "veces",
    "bien", "mal", "solo", "sólo", "así", "asi", "aquí", "aqui", "allí", "alli", "ahí", "ahi",
    "uno", "dos", "tres", "junta", "llega", "cosa", "cosas",
}

CV_DIR = None  # no aplica aquí

# Une cada centro de Clima Laboral con su tienda equivalente en el sistema de
# Reseñas (Google), para poder mostrar la satisfacción real de cliente junto
# al engagement. Oficinas no tiene tienda física de cara al público, así que
# se queda fuera (None) — igual que ParqueSur Fábrica: aunque comparte la
# misma ficha de Google que ParqueSur Tienda (una sola ubicación en Reseñas),
# la fábrica no atiende clientes, así que no tiene sentido mostrarle una
# satisfacción de cliente que en realidad es la de la tienda.
_CENTRO_A_TIENDAS = {
    "caleido": ["Caleido"],
    "gran plaza 2": ["Gran Plaza 2"],
    "la gavia": ["La Gavia"],
    "parquesur tienda": ["ParqueSur"],
    "princesa": ["Princesa"],
    "plenilunio": ["Plenilunio"],
}


def ensure_clima_tables():
    conn = get_connection()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS clima_oleadas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            etiqueta TEXT,
            creado_en TEXT NOT NULL DEFAULT (datetime('now'))
        )
    """)
    # empresa separa las oleadas de Krispy Kreme de las de Saona (mismo motor
    # de cálculo, datos completamente aparte). Las oleadas ya existentes son
    # todas de Krispy Kreme, así que el default cubre la migración solo.
    cols_oleadas = {row[1] for row in conn.execute("PRAGMA table_info(clima_oleadas)")}
    if "empresa" not in cols_oleadas:
        conn.execute("ALTER TABLE clima_oleadas ADD COLUMN empresa TEXT NOT NULL DEFAULT 'kk'")
    conn.execute("""
        CREATE TABLE IF NOT EXISTS clima_respuestas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            oleada_id INTEGER NOT NULL REFERENCES clima_oleadas(id),
            centro TEXT NOT NULL,
            fila_hash TEXT NOT NULL,
            datos_json TEXT NOT NULL,
            creado_en TEXT NOT NULL DEFAULT (datetime('now')),
            UNIQUE(oleada_id, fila_hash)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS clima_plantilla (
            oleada_id INTEGER NOT NULL REFERENCES clima_oleadas(id),
            centro TEXT NOT NULL,
            empleados INTEGER NOT NULL,
            PRIMARY KEY (oleada_id, centro)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS clima_importaciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            oleada_id INTEGER NOT NULL REFERENCES clima_oleadas(id),
            archivo_nombre TEXT,
            subido_por TEXT,
            subido_en TEXT NOT NULL DEFAULT (datetime('now')),
            nuevas INTEGER NOT NULL DEFAULT 0,
            ya_existian INTEGER NOT NULL DEFAULT 0
        )
    """)
    # Restricción por centro para un usuario dentro del módulo Clima Laboral
    # (p.ej. un gerente que solo debe ver el suyo) — tabla propia en vez de
    # reutilizar usuario_tiendas de Reseñas porque los nombres no coinciden
    # 1:1 (aquí ParqueSur se separa en "ParqueSur Fabrica"/"ParqueSur Tienda",
    # y existe "Oficinas", que no es una tienda física). Mismo patrón que
    # usuario_informe_tipos en informes.py: sin filas = sin restricción.
    conn.execute("""
        CREATE TABLE IF NOT EXISTS usuario_clima_centros (
            usuario_id INTEGER NOT NULL REFERENCES usuarios(id),
            centro TEXT NOT NULL,
            PRIMARY KEY (usuario_id, centro)
        )
    """)
    conn.commit()
    conn.close()


def get_centros_permitidos(usuario_id):
    """Centros de Clima Laboral a los que este usuario tiene acceso. Lista
    vacía = sin restricción (ve todos), igual que get_tiendas_permitidas en
    Reseñas y get_tipos_permitidos en Informes."""
    conn = get_connection()
    rows = conn.execute(
        "SELECT centro FROM usuario_clima_centros WHERE usuario_id = ? ORDER BY centro", (usuario_id,)
    ).fetchall()
    conn.close()
    return [r["centro"] for r in rows]


def set_centros_permitidos(usuario_id, centros):
    conn = get_connection()
    conn.execute("DELETE FROM usuario_clima_centros WHERE usuario_id = ?", (usuario_id,))
    for centro in centros:
        conn.execute(
            "INSERT OR IGNORE INTO usuario_clima_centros (usuario_id, centro) VALUES (?, ?)", (usuario_id, centro)
        )
    conn.commit()
    conn.close()


def _normaliza_header(h):
    return (h or "").strip().lower().replace("í", "i").replace("ó", "o").replace("á", "a").replace(
        "é", "e"
    ).replace("ú", "u")


def _hash_fila(fila):
    blob = json.dumps(fila, sort_keys=True, ensure_ascii=False, default=str)
    return hashlib.sha256(blob.encode("utf-8")).hexdigest()


def _read_sheet_rows(ws):
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        return []
    filas = []
    for row in rows_iter:
        if row is None or all(v is None for v in row):
            continue
        datos = {}
        for i, key in enumerate(header):
            if not key:
                continue
            value = row[i] if i < len(row) else None
            if hasattr(value, "isoformat"):
                value = value.isoformat()
            datos[key] = value
        if datos:
            filas.append(datos)
    return filas


# Formato posicional (Saona, Google Forms): sin categoría en el texto de la
# pregunta — el propio Apps Script que generaba estos informes a mano asume
# que, tras la columna de centro, vienen exactamente 5 preguntas de
# "Resultados de Engagement" + 20 de "Impulsores de Engagement", y las 2
# últimas columnas del formulario son siempre las preguntas abiertas.
POSICIONAL_N_RESULTADOS = 5
POSICIONAL_N_ABIERTAS = 2
POSICIONAL_CAT_RESULTADOS = "Resultados de Engagement"
POSICIONAL_CAT_IMPULSORES = "Impulsores de Engagement"


def _tiene_categoria_en_texto(h):
    """True solo si el "." separa una categoría real de una pregunta real
    (KK: "Resultados.Estoy contento...") — una frase que simplemente TERMINA
    en punto (frecuente en formularios sin esa convención) no cuenta: el
    partition dejaría la "pregunta" vacía."""
    if "." not in h:
        return False
    _, _, resto = h.partition(".")
    return len(resto.strip()) > 3


def _column_roles(headers):
    """Clasifica cada columna sin depender del texto exacto de la pregunta.
    Dos formatos reconocidos para las preguntas de escala:
    - Con categoría en el propio texto ("Categoria.Pregunta", Krispy Kreme):
      se usa tal cual, en el orden en que aparecen.
    - Sin categoría (Saona/Google Forms, preguntas como frase suelta): se
      asume por POSICIÓN — igual que el script que generaba estos informes
      a mano — las primeras 5 preguntas de escala son "Resultados de
      Engagement", el resto "Impulsores de Engagement", y las 2 últimas
      columnas del formulario son las preguntas abiertas.
    """
    centro_col = None
    metadata_cols = {}
    otras = []  # columnas que no son centro ni metadata, en orden

    for h in headers:
        norm = _normaliza_header(h)
        if "centro" in norm:
            centro_col = h
            continue
        matched_meta = False
        for hint, campo in METADATA_HINTS.items():
            if norm == hint:
                metadata_cols[campo] = h
                matched_meta = True
                break
        if matched_meta:
            continue
        otras.append(h)

    con_categoria = [h for h in otras if _tiene_categoria_en_texto(h)]

    if con_categoria:
        likert = []
        abiertas = []
        categorias_orden = []
        for h in otras:
            if _tiene_categoria_en_texto(h):
                categoria, _, pregunta = h.partition(".")
                categoria = categoria.strip()
                pregunta = pregunta.strip()
                if categoria not in categorias_orden:
                    categorias_orden.append(categoria)
                likert.append((h, categoria, pregunta))
            else:
                abiertas.append(h)
    else:
        if len(otras) > POSICIONAL_N_ABIERTAS:
            preguntas_escala = otras[:-POSICIONAL_N_ABIERTAS]
            abiertas = otras[-POSICIONAL_N_ABIERTAS:]
        else:
            preguntas_escala = otras
            abiertas = []
        likert = [
            (h, POSICIONAL_CAT_RESULTADOS if i < POSICIONAL_N_RESULTADOS else POSICIONAL_CAT_IMPULSORES, h)
            for i, h in enumerate(preguntas_escala)
        ]
        categorias_orden = [POSICIONAL_CAT_RESULTADOS, POSICIONAL_CAT_IMPULSORES] if preguntas_escala else []

    return {
        "centro_col": centro_col,
        "metadata": metadata_cols,
        "likert": likert,
        "abiertas": abiertas,
        "categorias_orden": categorias_orden,
    }


def _parse_plantilla(wb):
    if "Plantilla" not in wb.sheetnames:
        return {}
    ws = wb["Plantilla"]
    filas = _read_sheet_rows(ws)
    resultado = {}
    for fila in filas:
        centro = None
        empleados = None
        for k, v in fila.items():
            kn = _normaliza_header(k)
            if kn == "tienda" or "centro" in kn:
                centro = v
            elif "emplead" in kn:
                empleados = v
        if centro is None or empleados is None:
            # Algunas plantillas (Saona) no usan estos nombres de columna en
            # absoluto (p.ej. "Nombre"/"Sede" en vez de "Centro"/"Empleados")
            # — igual que el script que generaba estos informes a mano, se
            # cae a POSICIÓN: primera columna = centro, segunda = empleados.
            valores = list(fila.values())
            if len(valores) >= 2:
                centro = centro if centro is not None else valores[0]
                empleados = empleados if empleados is not None else valores[1]
        if centro and empleados is not None and str(centro).strip().lower() != "total":
            try:
                resultado[str(centro).strip()] = int(empleados)
            except (TypeError, ValueError):
                continue
    return resultado


def get_or_create_oleada(nueva, etiqueta=None, empresa="kk"):
    conn = get_connection()
    if not nueva:
        row = conn.execute(
            "SELECT id FROM clima_oleadas WHERE empresa = ? ORDER BY id DESC LIMIT 1", (empresa,)
        ).fetchone()
        if row:
            oleada_id = row[0]
            conn.close()
            return oleada_id
    cur = conn.execute("INSERT INTO clima_oleadas (etiqueta, empresa) VALUES (?, ?)", (etiqueta, empresa))
    conn.commit()
    oleada_id = cur.lastrowid
    conn.close()
    return oleada_id


def get_oleada_empresa(oleada_id):
    conn = get_connection()
    row = conn.execute("SELECT empresa FROM clima_oleadas WHERE id = ?", (oleada_id,)).fetchone()
    conn.close()
    return row["empresa"] if row else None


def import_excel(file_bytes, archivo_nombre, subido_por, nueva_oleada=False, empresa="kk"):
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as e:
        raise ValueError(f"No se pudo leer el archivo Excel: {e}")

    if "Respuestas" not in wb.sheetnames:
        raise ValueError("El Excel debe tener una hoja llamada 'Respuestas'")

    filas = _read_sheet_rows(wb["Respuestas"])
    if not filas:
        raise ValueError("La hoja 'Respuestas' no tiene filas de datos")

    roles = _column_roles(list(filas[0].keys()))
    if roles["centro_col"] is None:
        raise ValueError("No se encontró una columna de centro de trabajo (p.ej. '¿Cuál es tu centro de trabajo?')")

    plantilla = _parse_plantilla(wb)

    oleada_id = get_or_create_oleada(nueva_oleada, empresa=empresa)

    conn = get_connection()
    cur = conn.execute(
        "INSERT INTO clima_importaciones (oleada_id, archivo_nombre, subido_por) VALUES (?, ?, ?)",
        (oleada_id, archivo_nombre, subido_por),
    )
    importacion_id = cur.lastrowid

    nuevas = 0
    ya_existian = 0
    for fila in filas:
        centro = fila.get(roles["centro_col"])
        if not centro:
            continue
        fila_hash = _hash_fila(fila)
        existe = conn.execute(
            "SELECT id FROM clima_respuestas WHERE oleada_id = ? AND fila_hash = ?", (oleada_id, fila_hash)
        ).fetchone()
        if existe:
            ya_existian += 1
            continue
        datos_json = json.dumps(fila, ensure_ascii=False, default=str)
        conn.execute(
            "INSERT INTO clima_respuestas (oleada_id, centro, fila_hash, datos_json) VALUES (?, ?, ?, ?)",
            (oleada_id, str(centro).strip(), fila_hash, datos_json),
        )
        nuevas += 1

    for centro, empleados in plantilla.items():
        conn.execute(
            "INSERT INTO clima_plantilla (oleada_id, centro, empleados) VALUES (?, ?, ?) "
            "ON CONFLICT(oleada_id, centro) DO UPDATE SET empleados = excluded.empleados",
            (oleada_id, centro, empleados),
        )

    conn.execute(
        "UPDATE clima_importaciones SET nuevas = ?, ya_existian = ? WHERE id = ?",
        (nuevas, ya_existian, importacion_id),
    )
    conn.commit()
    conn.close()
    return {"oleada_id": oleada_id, "nuevas": nuevas, "ya_existian": ya_existian, "total_en_excel": len(filas)}


def list_oleadas(empresa="kk"):
    # numero: posición de esta oleada DENTRO de su propia empresa (1, 2, 3...)
    # — el id de clima_oleadas es un autoincrement GLOBAL compartido entre
    # KK y Saona, así que la primera oleada de Saona puede tener un id como
    # 3 aunque sea la primera que se sube; "numero" evita que eso confunda
    # en el selector ("Oleada #3" cuando en realidad es la primera de Saona).
    conn = get_connection()
    rows = conn.execute("""
        SELECT o.id, o.etiqueta, o.creado_en, COUNT(r.id) AS num_respuestas,
               ROW_NUMBER() OVER (ORDER BY o.id ASC) AS numero
        FROM clima_oleadas o
        LEFT JOIN clima_respuestas r ON r.oleada_id = o.id
        WHERE o.empresa = ?
        GROUP BY o.id ORDER BY o.id DESC
    """, (empresa,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def list_centros_conocidos():
    """Todos los centros que han aparecido alguna vez en cualquier oleada (de
    cualquier empresa) — a diferencia de Informes, Clima no tiene un catálogo
    fijo de tipos, así que el checklist de restricción por centro en Usuarios
    se arma a partir de lo que ya se ha importado."""
    conn = get_connection()
    rows = conn.execute("SELECT DISTINCT centro FROM clima_respuestas ORDER BY centro").fetchall()
    conn.close()
    return [r["centro"] for r in rows]


def list_centros(oleada_id, centros_permitidos=None):
    conn = get_connection()
    rows = conn.execute(
        "SELECT DISTINCT centro FROM clima_respuestas WHERE oleada_id = ? ORDER BY centro", (oleada_id,)
    ).fetchall()
    conn.close()
    centros = [r["centro"] for r in rows]
    if centros_permitidos:
        centros = [c for c in centros if c in centros_permitidos]
    return centros


def _es_oficinas(centro):
    return _normaliza_header(centro) == "oficinas"


def _es_fabrica(centro):
    """Detecta cualquier centro cuyo nombre contenga "fábrica" (ParqueSur
    Fábrica hoy, y las que se vayan añadiendo) — por nombre y no por una
    lista fija, para que una fábrica nueva aparezca sola en su gráfico sin
    tener que tocar código cada vez."""
    return "fabrica" in _normaliza_header(centro)


def _fetch_respuestas(conn, oleada_id, centro, centros_permitidos=None):
    if centro:
        rows = conn.execute(
            "SELECT datos_json FROM clima_respuestas WHERE oleada_id = ? AND centro = ?", (oleada_id, centro)
        ).fetchall()
        return [json.loads(r["datos_json"]) for r in rows]
    # Vista "todos los centros": Oficinas no es una tienda (no tiene cliente
    # ni ventas) y distorsionaría el engagement/participación agregados de
    # las tiendas — se excluye aquí, pero sigue siendo su propio centro
    # seleccionable individualmente en el grid. centros_permitidos (gerente
    # restringido a su/s centro/s, ver usuario_clima_centros) se aplica igual
    # que la exclusión de Oficinas, para que el agregado de "todos los
    # centros" de un gerente restringido solo cuente los suyos.
    rows = conn.execute("SELECT datos_json, centro FROM clima_respuestas WHERE oleada_id = ?", (oleada_id,)).fetchall()
    return [
        json.loads(r["datos_json"]) for r in rows
        if not _es_oficinas(r["centro"]) and (not centros_permitidos or r["centro"] in centros_permitidos)
    ]


def _empleados_total(conn, oleada_id, centro, centros_permitidos=None):
    if centro:
        row = conn.execute(
            "SELECT empleados FROM clima_plantilla WHERE oleada_id = ? AND centro = ?", (oleada_id, centro)
        ).fetchone()
        return row["empleados"] if row else None
    rows = conn.execute("SELECT centro, empleados FROM clima_plantilla WHERE oleada_id = ?", (oleada_id,)).fetchall()
    if not rows:
        return None
    return sum(
        r["empleados"] for r in rows
        if not _es_oficinas(r["centro"]) and (not centros_permitidos or r["centro"] in centros_permitidos)
    )


def _tokenizar(texto):
    palabras = re.findall(r"[a-záéíóúñü]+", texto.lower())
    return [p for p in palabras if len(p) > 2 and p not in STOPWORDS]


_RESPUESTAS_SIN_VALOR = {".", "..", "…", "-", "sin comentarios", "sin comentario"}


def _respuesta_con_valor(texto):
    """Filtra respuestas abiertas vacías o que son solo un punto/relleno
    ("." , "sin comentarios") — no aportan nada y no deben aparecer en la
    lista de comentarios ni en el PDF."""
    limpio = (texto or "").strip()
    if not limpio:
        return False
    return limpio.lower() not in _RESPUESTAS_SIN_VALOR


def compute_reporte(oleada_id, centro=None, centros_permitidos=None):
    conn = get_connection()
    filas = _fetch_respuestas(conn, oleada_id, centro, centros_permitidos)
    if not filas:
        conn.close()
        raise ValueError("No hay respuestas para este centro en esta oleada")

    roles = _column_roles(list(filas[0].keys()))
    n = len(filas)
    empleados = _empleados_total(conn, oleada_id, centro, centros_permitidos)
    participacion = round(n / empleados * 100, 1) if empleados else None

    def stats_pregunta(header):
        conteo = {cat: 0 for cat in LIKERT_ORDEN}
        respondidas = 0
        for fila in filas:
            valor = _canonicalizar_likert(fila.get(header))
            if valor in conteo:
                conteo[valor] += 1
                respondidas += 1
        if respondidas == 0:
            porcentajes = {cat: 0.0 for cat in LIKERT_ORDEN}
        else:
            porcentajes = {cat: round(conteo[cat] / respondidas * 100, 1) for cat in LIKERT_ORDEN}
        top2box = round(porcentajes["Totalmente de acuerdo"] + porcentajes["De acuerdo"], 1)
        # "Oportunidad" no es una pareja fija de categorías: es la suma de
        # las DOS que más pesen entre Neutral / En desacuerdo / Totalmente en
        # desacuerdo (la tercera, la que menos pese de esas tres, queda
        # fuera). Con "Totalmente en desacuerdo" en 0%, por ejemplo, las dos
        # que cuentan pueden perfectamente ser Neutral + En desacuerdo — no
        # siempre son las mismas dos categorías.
        negativas = sorted(
            (porcentajes[c] for c in ("Neutral", "En desacuerdo", "Totalmente en desacuerdo")),
            reverse=True,
        )
        oportunidad = round(negativas[0] + negativas[1], 1)
        return porcentajes, top2box, oportunidad

    primera_categoria = roles["categorias_orden"][0] if roles["categorias_orden"] else None

    resultados_engagement = []
    impulsores_engagement = []
    todas_preguntas_top2box = []

    for header, categoria, pregunta in roles["likert"]:
        porcentajes, top2box, oportunidad = stats_pregunta(header)
        item = {
            "pregunta": pregunta,
            "categoria": categoria,
            "porcentajes": porcentajes,
            "top2box": top2box,
            "oportunidad": oportunidad,
        }
        if categoria == primera_categoria:
            resultados_engagement.append(item)
        else:
            impulsores_engagement.append(item)
        todas_preguntas_top2box.append(item)

    if resultados_engagement:
        engagement_score = round(
            sum(i["porcentajes"]["Totalmente de acuerdo"] for i in resultados_engagement) / len(resultados_engagement),
            1,
        )
    else:
        engagement_score = None

    # Mismo criterio que se muestra en pantalla: fortalezas = mayor top2box
    # (Totalmente de acuerdo + De acuerdo); oportunidades = mayor
    # "oportunidad" (las dos categorías que más pesen entre Neutral/En
    # desacuerdo/Totalmente en desacuerdo — ver stats_pregunta). Empates se
    # resuelven por la casilla más extrema de cada lado.
    fortalezas = sorted(
        todas_preguntas_top2box,
        key=lambda i: (i["top2box"], i["porcentajes"]["Totalmente de acuerdo"]),
        reverse=True,
    )[:2]
    oportunidades = sorted(
        todas_preguntas_top2box,
        key=lambda i: (i["oportunidad"], i["porcentajes"]["Totalmente en desacuerdo"]),
        reverse=True,
    )[:2]

    abiertas = {}
    nube_palabras = {}
    for header in roles["abiertas"]:
        textos = [str(fila[header]).strip() for fila in filas if fila.get(header)]
        textos = [t for t in textos if _respuesta_con_valor(t)]
        abiertas[header] = textos
        conteo_palabras = {}
        for texto in textos:
            for palabra in _tokenizar(texto):
                conteo_palabras[palabra] = conteo_palabras.get(palabra, 0) + 1
        top_palabras = sorted(conteo_palabras.items(), key=lambda x: x[1], reverse=True)[:40]
        nube_palabras[header] = [{"palabra": p, "veces": c} for p, c in top_palabras]

    anterior = get_anterior_score(centro, oleada_id)
    satisfaccion = get_satisfaccion_cliente(centro, oleada_id)

    conn.close()
    return {
        "oleada_id": oleada_id,
        "centro": centro,
        "n": n,
        "empleados": empleados,
        "participacion": participacion,
        "engagement_presente": engagement_score,
        "engagement_anterior": anterior,
        "satisfaccion_presente": satisfaccion["presente"] if satisfaccion else None,
        "satisfaccion_anterior": satisfaccion["anterior"] if satisfaccion else None,
        "tiene_satisfaccion": satisfaccion is not None,
        "resultados_engagement": resultados_engagement,
        "impulsores_engagement": impulsores_engagement,
        "fortalezas": fortalezas,
        "oportunidades": oportunidades,
        "abiertas": abiertas,
        "nube_palabras": nube_palabras,
    }


def compute_por_centro(oleada_id, centros_permitidos=None):
    """Engagement y participación de CADA tienda por separado (Oficinas
    excluida, igual que en el agregado) — alimenta los gráficos de barras de
    la vista "todos los centros". Separado en tiendas/fábricas porque son
    realidades distintas (una fábrica no tiene el mismo perfil de plantilla
    ni de participación que una tienda) y mezclarlas en el mismo gráfico de
    barras no deja comparar bien ni unas ni otras."""
    centros = [c for c in list_centros(oleada_id, centros_permitidos) if not _es_oficinas(c)]
    tiendas = []
    fabricas = []
    for centro in centros:
        try:
            reporte = compute_reporte(oleada_id, centro)
        except ValueError:
            continue
        item = {
            "centro": centro,
            "engagement": reporte["engagement_presente"],
            "participacion": reporte["participacion"],
        }
        (fabricas if _es_fabrica(centro) else tiendas).append(item)
    return {"tiendas": tiendas, "fabricas": fabricas}


def _tiendas_para_centro(centro, empresa="kk"):
    # _CENTRO_A_TIENDAS solo mapea centros de Krispy Kreme a su ficha de
    # Google (Saona todavía no tiene su propio sistema de Reseñas conectado,
    # ver tarea de scraping) — así que para Saona nunca hay con qué comparar
    # satisfacción de cliente, y no se debe caer en la lista de tiendas KK.
    if empresa != "kk":
        return None
    if centro is None:
        todas = set()
        for tiendas in _CENTRO_A_TIENDAS.values():
            todas.update(tiendas)
        return sorted(todas)
    return _CENTRO_A_TIENDAS.get(_normaliza_header(centro))


def get_satisfaccion_cliente(centro, oleada_id):
    """Satisfacción de cliente (estrellas de Google) para el mismo centro,
    para mostrarla junto al engagement. Presente = media de todas las
    reseñas a día de hoy. Anterior = media de reseñas hasta la fecha de la
    oleada anterior (así se compara el mismo punto en el tiempo que el
    engagement de esa oleada). None si el centro no tiene tienda física
    (p.ej. Oficinas), no hay reseñas, o la oleada es de otra empresa sin
    Reseñas propias conectadas."""
    empresa = get_oleada_empresa(oleada_id) or "kk"
    tiendas = _tiendas_para_centro(centro, empresa)
    if not tiendas:
        return None

    conn = get_connection()
    # AND empresa = ?: la oleada "anterior" tiene que ser de la MISMA
    # empresa — si no, una Saona compararía su engagement contra el de
    # Krispy Kreme solo porque su id es menor (bug real que se dio con las
    # primeras oleadas de Saona).
    anterior_oleada = conn.execute(
        "SELECT creado_en FROM clima_oleadas WHERE id < ? AND empresa = ? ORDER BY id DESC LIMIT 1",
        (oleada_id, empresa),
    ).fetchone()
    fecha_anterior = anterior_oleada["creado_en"] if anterior_oleada else None

    placeholders = ",".join("?" for _ in tiendas)

    def promedio(fecha_limite):
        query = f"SELECT AVG(calificacion_num) AS media FROM reviews WHERE tienda IN ({placeholders})"
        params = list(tiendas)
        if fecha_limite:
            query += " AND fecha_datetime <= ?"
            params.append(fecha_limite)
        row = conn.execute(query, params).fetchone()
        return round(row["media"], 2) if row and row["media"] is not None else None

    presente = promedio(None)
    anterior = promedio(fecha_anterior) if fecha_anterior else None
    conn.close()
    return {"presente": presente, "anterior": anterior}


def get_anterior_score(centro, oleada_id):
    # Igual que en get_satisfaccion_cliente: la "oleada anterior" tiene que
    # ser de la MISMA empresa, si no una oleada de Saona podría tomar el
    # engagement de Krispy Kreme como su "Anterior" solo por tener un id
    # más alto (bug real reportado con las primeras oleadas de Saona).
    empresa = get_oleada_empresa(oleada_id) or "kk"
    conn = get_connection()
    oleadas_previas = conn.execute(
        "SELECT id FROM clima_oleadas WHERE id < ? AND empresa = ? ORDER BY id DESC", (oleada_id, empresa)
    ).fetchall()
    conn.close()
    for row in oleadas_previas:
        try:
            reporte = compute_reporte(row["id"], centro)
        except ValueError:
            continue
        if reporte["engagement_presente"] is not None:
            return reporte["engagement_presente"]
    return None


ensure_clima_tables()
