import datetime
import hashlib
import io
import json
import os

from openpyxl import load_workbook

import reclutamiento as reclutamiento_module
import scoring_valores
from db import DATA_DIR, get_connection

# Estos son los que ya sabemos que existen; el admin puede añadir más desde
# la pantalla de Informes a medida que surjan nuevas encuestas de Forms.
# Entrevistas de Salida tiene su propio módulo dedicado (backend/entrevistas.py,
# igual que Clima Laboral) — no vive aquí.
DEFAULT_TIPOS = [
    ("valores_oficina", "Valores y Competencias — Oficina"),
    ("valores_tiendas", "Valores y Competencias — Tiendas"),
]

# Equivalentes de Saona — mismo sistema genérico, separados solo por
# "empresa" en informe_tipos. El prefijo "SAONA · " en el nombre es para que
# se distingan en el checklist de permisos de Usuarios, donde tipos de las
# dos empresas aparecen mezclados en una sola lista. Igual que Krispy Kreme,
# Valores y Competencias se separa en Restaurantes/Oficina (dos poblaciones
# distintas, no tiene sentido mezclarlas en un solo tipo).
DEFAULT_TIPOS_SAONA = [
    ("saona_valores_restaurantes", "SAONA · Valores y Competencias — Restaurantes"),
    ("saona_valores_oficina", "SAONA · Valores y Competencias — Oficina"),
]

CV_DIR = os.path.join(DATA_DIR, "uploads", "cv")

DATE_HINTS = ("fecha", "hora", "date")


def ensure_informe_tables():
    conn = get_connection()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS informe_tipos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            clave TEXT UNIQUE NOT NULL,
            nombre TEXT NOT NULL,
            creado TEXT NOT NULL DEFAULT (datetime('now'))
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS informe_importaciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tipo_id INTEGER NOT NULL REFERENCES informe_tipos(id),
            archivo_nombre TEXT,
            subido_en TEXT NOT NULL DEFAULT (datetime('now')),
            subido_por TEXT,
            num_respuestas INTEGER NOT NULL DEFAULT 0
        )
    """)
    # datos_json guarda la fila completa {pregunta: respuesta} tal cual viene
    # del Excel: cada formulario/hoja tiene sus propias columnas, así que no
    # hay un esquema fijo posible aquí. El dedup usa un hash de la fila
    # completa (no un ID de formulario, que muchos de estos Excel no traen).
    conn.execute("""
        CREATE TABLE IF NOT EXISTS informe_respuestas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tipo_id INTEGER NOT NULL REFERENCES informe_tipos(id),
            importacion_id INTEGER NOT NULL REFERENCES informe_importaciones(id),
            hoja TEXT NOT NULL DEFAULT 'Scoring',
            fila_hash TEXT NOT NULL,
            datos_json TEXT NOT NULL,
            cv_ruta TEXT,
            cv_nombre_original TEXT,
            creado_en TEXT NOT NULL DEFAULT (datetime('now')),
            UNIQUE(tipo_id, hoja, fila_hash)
        )
    """)
    # Sin filas = ve todos los tipos de informe (igual que usuario_tiendas en
    # Reseñas) — esto es un refinamiento DENTRO del módulo "informes", que ya
    # se concede o no a nivel de módulo en usuario_modulos.
    conn.execute("""
        CREATE TABLE IF NOT EXISTS usuario_informe_tipos (
            usuario_id INTEGER NOT NULL REFERENCES usuarios(id),
            tipo_clave TEXT NOT NULL,
            PRIMARY KEY (usuario_id, tipo_clave)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS informe_compartidos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            respuesta_id INTEGER NOT NULL REFERENCES informe_respuestas(id),
            usuario_id INTEGER NOT NULL REFERENCES usuarios(id),
            compartido_por TEXT,
            compartido_en TEXT NOT NULL DEFAULT (datetime('now')),
            UNIQUE(respuesta_id, usuario_id)
        )
    """)
    # Algunos Excel traen hojas auxiliares (ayudas de fórmulas, respuestas
    # crudas antes de procesar, etc.) que no son vistas útiles para el
    # usuario — se pueden ocultar sin borrar los datos.
    conn.execute("""
        CREATE TABLE IF NOT EXISTS informe_hojas_ocultas (
            tipo_id INTEGER NOT NULL REFERENCES informe_tipos(id),
            hoja TEXT NOT NULL,
            PRIMARY KEY (tipo_id, hoja)
        )
    """)

    # hoja_conteo: qué hoja representa el conteo "real" de respuestas del
    # tipo cuando varias hojas son distintas vistas de las MISMAS personas
    # (p.ej. Scoring y Dashboard) — evita sumar y duplicar el total.
    cols_tipos = {row[1] for row in conn.execute("PRAGMA table_info(informe_tipos)")}
    if "hoja_conteo" not in cols_tipos:
        conn.execute("ALTER TABLE informe_tipos ADD COLUMN hoja_conteo TEXT")
    # empresa separa los tipos de Krispy Kreme de los de Saona (mismo sistema
    # genérico de Informes, datos completamente aparte). Los tipos ya
    # existentes son todos de Krispy Kreme, de ahí el default.
    if "empresa" not in cols_tipos:
        conn.execute("ALTER TABLE informe_tipos ADD COLUMN empresa TEXT NOT NULL DEFAULT 'kk'")

    # candidato_id enlaza cada "compartido" con su ficha en candidatos
    # (backend/reclutamiento.py) — se rellena al compartir (ver
    # compartir_respuestas). Las filas de antes de esta columna se resuelven
    # también en compartir_respuestas la próxima vez que se vuelvan a
    # compartir; mientras tanto quedan como NULL sin romper nada.
    cols_compartidos = {row[1] for row in conn.execute("PRAGMA table_info(informe_compartidos)")}
    if "candidato_id" not in cols_compartidos:
        conn.execute("ALTER TABLE informe_compartidos ADD COLUMN candidato_id INTEGER REFERENCES candidatos(id)")

    _migrate_legacy_respuestas(conn)

    # Renombrado explícito: "Operativa" pasó a llamarse "Tiendas" (mismo id,
    # así que las respuestas ya importadas no se ven afectadas). Va ANTES del
    # INSERT OR IGNORE de abajo para que no choquen las claves.
    conn.execute("""
        UPDATE informe_tipos SET clave = 'valores_tiendas', nombre = 'Valores y Competencias — Tiendas'
        WHERE clave = 'valores_operativa'
    """)

    # Mismo caso para Saona: el primer tipo de Valores y Competencias se creó
    # sin distinguir Restaurantes/Oficina; al separarlo se reutiliza esa
    # misma fila (sin respuestas todavía) para "Restaurantes" y se añade
    # "Oficina" como tipo nuevo más abajo.
    conn.execute("""
        UPDATE informe_tipos SET clave = 'saona_valores_restaurantes',
            nombre = 'SAONA · Valores y Competencias — Restaurantes'
        WHERE clave = 'saona_valores'
    """)

    # Clima Laboral pasó a tener su propio módulo dedicado (backend/clima.py)
    # — no encaja en el modelo genérico fila=candidato. Se quita de aquí solo
    # si nunca se llegó a usar (0 respuestas), para no perder datos reales.
    conn.execute("""
        DELETE FROM informe_tipos WHERE clave = 'clima_laboral'
          AND id NOT IN (SELECT DISTINCT tipo_id FROM informe_respuestas)
    """)
    # Mismo caso: Entrevistas de Salida pasó a tener su propio módulo
    # dedicado (backend/entrevistas.py), igual que Clima Laboral. A
    # diferencia de Clima, aquí se borra aunque tenga respuestas: es un
    # volcado genérico obsoleto (el usuario ya subió el Excel real al módulo
    # dedicado), así que se elimina en cascada para que la tarjeta desaparezca
    # de Informes.
    conn.execute("""
        DELETE FROM informe_compartidos WHERE respuesta_id IN (
            SELECT id FROM informe_respuestas WHERE tipo_id IN (
                SELECT id FROM informe_tipos WHERE clave IN ('entrevistas_salida', 'saona_entrevistas_salida')
            )
        )
    """)
    conn.execute("""
        DELETE FROM informe_respuestas WHERE tipo_id IN (
            SELECT id FROM informe_tipos WHERE clave IN ('entrevistas_salida', 'saona_entrevistas_salida')
        )
    """)
    conn.execute("""
        DELETE FROM informe_importaciones WHERE tipo_id IN (
            SELECT id FROM informe_tipos WHERE clave IN ('entrevistas_salida', 'saona_entrevistas_salida')
        )
    """)
    conn.execute("""
        DELETE FROM informe_hojas_ocultas WHERE tipo_id IN (
            SELECT id FROM informe_tipos WHERE clave IN ('entrevistas_salida', 'saona_entrevistas_salida')
        )
    """)
    conn.execute("""
        DELETE FROM usuario_informe_tipos WHERE tipo_clave IN ('entrevistas_salida', 'saona_entrevistas_salida')
    """)
    conn.execute("""
        DELETE FROM informe_tipos WHERE clave IN ('entrevistas_salida', 'saona_entrevistas_salida')
    """)

    for clave, nombre in DEFAULT_TIPOS:
        conn.execute("INSERT OR IGNORE INTO informe_tipos (clave, nombre, empresa) VALUES (?, ?, 'kk')", (clave, nombre))
    for clave, nombre in DEFAULT_TIPOS_SAONA:
        conn.execute("INSERT OR IGNORE INTO informe_tipos (clave, nombre, empresa) VALUES (?, ?, 'saona')", (clave, nombre))

    conn.commit()
    conn.close()
    os.makedirs(CV_DIR, exist_ok=True)


def _migrate_legacy_respuestas(conn):
    """La tabla original no tenía hoja/fila_hash/cv_* ni la UNIQUE nueva —
    si existe con el esquema viejo, se recrea y se rellenan esas columnas
    para las filas ya importadas (hoja='Scoring', fila_hash calculado sobre
    su propio datos_json)."""
    cols = {row[1] for row in conn.execute("PRAGMA table_info(informe_respuestas)")}
    if not cols or "fila_hash" in cols:
        return

    conn.execute("ALTER TABLE informe_respuestas RENAME TO informe_respuestas_old")
    conn.execute("""
        CREATE TABLE informe_respuestas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tipo_id INTEGER NOT NULL REFERENCES informe_tipos(id),
            importacion_id INTEGER NOT NULL REFERENCES informe_importaciones(id),
            hoja TEXT NOT NULL DEFAULT 'Scoring',
            fila_hash TEXT NOT NULL,
            datos_json TEXT NOT NULL,
            cv_ruta TEXT,
            cv_nombre_original TEXT,
            creado_en TEXT NOT NULL DEFAULT (datetime('now')),
            UNIQUE(tipo_id, hoja, fila_hash)
        )
    """)
    filas_viejas = conn.execute(
        "SELECT id, tipo_id, importacion_id, datos_json, creado_en FROM informe_respuestas_old"
    ).fetchall()
    for fila in filas_viejas:
        fila_hash = hash_fila(json.loads(fila["datos_json"]))
        conn.execute(
            """INSERT INTO informe_respuestas
               (id, tipo_id, importacion_id, hoja, fila_hash, datos_json, creado_en)
               VALUES (?, ?, ?, 'Scoring', ?, ?, ?)
               ON CONFLICT(tipo_id, hoja, fila_hash) DO NOTHING""",
            (fila["id"], fila["tipo_id"], fila["importacion_id"], fila_hash, fila["datos_json"], fila["creado_en"]),
        )
    conn.execute("DROP TABLE informe_respuestas_old")


def hash_fila(fila: dict) -> str:
    blob = json.dumps(fila, sort_keys=True, ensure_ascii=False, default=str)
    return hashlib.sha256(blob.encode("utf-8")).hexdigest()


def get_tipo(clave):
    conn = get_connection()
    row = conn.execute("SELECT * FROM informe_tipos WHERE clave = ?", (clave,)).fetchone()
    conn.close()
    return dict(row) if row else None


def _hoja_para_conteo(conn, tipo):
    """Cuando un tipo tiene varias hojas que son distintas vistas de las
    MISMAS personas (p.ej. Scoring y Dashboard), sumar sus filas duplicaría
    el total. Por eso el conteo del tipo se basa en UNA sola hoja: la que el
    admin eligió (hoja_conteo) o, si no ha elegido, la primera hoja visible
    en orden alfabético (mismo criterio que el hoja por defecto al ver
    respuestas)."""
    if tipo.get("hoja_conteo"):
        return tipo["hoja_conteo"]
    row = conn.execute("""
        SELECT hoja FROM informe_respuestas
        WHERE tipo_id = ? AND hoja NOT IN (SELECT hoja FROM informe_hojas_ocultas WHERE tipo_id = ?)
        GROUP BY hoja ORDER BY hoja LIMIT 1
    """, (tipo["id"], tipo["id"])).fetchone()
    return row["hoja"] if row else None


def list_tipos(empresa=None):
    conn = get_connection()
    if empresa:
        tipos = [dict(r) for r in conn.execute(
            "SELECT * FROM informe_tipos WHERE empresa = ? ORDER BY id", (empresa,)
        ).fetchall()]
    else:
        tipos = [dict(r) for r in conn.execute("SELECT * FROM informe_tipos ORDER BY id").fetchall()]
    for tipo in tipos:
        hoja = _hoja_para_conteo(conn, tipo)
        if hoja is None:
            tipo["num_respuestas"] = 0
        else:
            tipo["num_respuestas"] = conn.execute(
                "SELECT COUNT(*) FROM informe_respuestas WHERE tipo_id = ? AND hoja = ?", (tipo["id"], hoja)
            ).fetchone()[0]
        tipo["hoja_conteo_actual"] = hoja
    conn.close()
    return tipos


def create_tipo(clave, nombre, empresa="kk"):
    conn = get_connection()
    cur = conn.execute("INSERT INTO informe_tipos (clave, nombre, empresa) VALUES (?, ?, ?)", (clave, nombre, empresa))
    conn.commit()
    tipo_id = cur.lastrowid
    conn.close()
    return tipo_id


def list_hojas(tipo_clave, incluir_ocultas=True):
    tipo = get_tipo(tipo_clave)
    if tipo is None:
        raise ValueError(f"Tipo de informe desconocido: {tipo_clave}")
    conn = get_connection()
    rows = conn.execute(
        "SELECT hoja, COUNT(*) AS total FROM informe_respuestas WHERE tipo_id = ? GROUP BY hoja ORDER BY hoja",
        (tipo["id"],),
    ).fetchall()
    ocultas = {
        r[0] for r in conn.execute("SELECT hoja FROM informe_hojas_ocultas WHERE tipo_id = ?", (tipo["id"],))
    }
    hoja_conteo = _hoja_para_conteo(conn, tipo)
    conn.close()
    resultado = [
        {"hoja": r["hoja"], "total": r["total"], "oculta": r["hoja"] in ocultas, "principal": r["hoja"] == hoja_conteo}
        for r in rows
    ]
    if not incluir_ocultas:
        resultado = [r for r in resultado if not r["oculta"]]
    return resultado


def set_hoja_oculta(tipo_clave, hoja, oculta):
    tipo = get_tipo(tipo_clave)
    if tipo is None:
        raise ValueError(f"Tipo de informe desconocido: {tipo_clave}")
    conn = get_connection()
    if oculta:
        conn.execute(
            "INSERT OR IGNORE INTO informe_hojas_ocultas (tipo_id, hoja) VALUES (?, ?)", (tipo["id"], hoja)
        )
    else:
        conn.execute(
            "DELETE FROM informe_hojas_ocultas WHERE tipo_id = ? AND hoja = ?", (tipo["id"], hoja)
        )
    conn.commit()
    conn.close()


def set_hoja_conteo(tipo_clave, hoja):
    tipo = get_tipo(tipo_clave)
    if tipo is None:
        raise ValueError(f"Tipo de informe desconocido: {tipo_clave}")
    conn = get_connection()
    conn.execute("UPDATE informe_tipos SET hoja_conteo = ? WHERE id = ?", (hoja, tipo["id"]))
    conn.commit()
    conn.close()


def eliminar_hoja(tipo_clave, hoja):
    """Borra por completo una hoja que no es un dato real (p.ej. una pestaña
    auxiliar/script colada por error) — a diferencia de ocultar, esto sí
    elimina las filas."""
    tipo = get_tipo(tipo_clave)
    if tipo is None:
        raise ValueError(f"Tipo de informe desconocido: {tipo_clave}")
    conn = get_connection()
    conn.execute("""
        DELETE FROM informe_compartidos WHERE respuesta_id IN (
            SELECT id FROM informe_respuestas WHERE tipo_id = ? AND hoja = ?
        )
    """, (tipo["id"], hoja))
    conn.execute("DELETE FROM informe_respuestas WHERE tipo_id = ? AND hoja = ?", (tipo["id"], hoja))
    conn.execute("DELETE FROM informe_hojas_ocultas WHERE tipo_id = ? AND hoja = ?", (tipo["id"], hoja))
    if tipo.get("hoja_conteo") == hoja:
        conn.execute("UPDATE informe_tipos SET hoja_conteo = NULL WHERE id = ?", (tipo["id"],))
    conn.commit()
    conn.close()


def read_workbook_sheets(file_bytes):
    """Lee TODAS las hojas VISIBLES con datos tabulares del Excel (primera
    fila = cabeceras, resto = filas). Hojas ocultas/muy ocultas en el propio
    Excel (auxiliares, cálculos internos) y hojas sin filas de datos (p.ej.
    solo gráficos) se ignoran. No asume qué columnas existen: cada
    hoja/formulario trae las suyas."""
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"No se pudo leer el archivo Excel: {e}")

    resultado = {}
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
        except StopIteration:
            continue
        if not any(header):
            continue

        filas = []
        for row in rows_iter:
            if row is None or all(v is None for v in row):
                continue
            datos = {}
            for i, key in enumerate(header):
                if not key:
                    continue
                value = row[i] if i < len(row) else None
                if hasattr(value, "isoformat"):
                    value = value.isoformat()
                datos[key] = value
            if datos:
                filas.append(datos)
        if filas:
            resultado[ws.title] = filas
    return resultado


def _quiza_calcular_scoring(hojas, columnas_extra=None):
    """Si el Excel trae una hoja "Respuestas" cruda (export directo de Forms,
    sin Dashboard ya calculado) y se reconoce como de Valores y Competencias
    por sus preguntas, calculamos aquí el Scoring/Dashboard — así el usuario
    solo sube el Excel plano sin depender del script externo. Si el Excel ya
    trae su propio Dashboard (el flujo de siempre), no se toca nada. Funciona
    igual para cualquier tipo/empresa (Krispy Kreme, Saona...) porque detecta
    el cuestionario por su contenido, no por el nombre del tipo.

    columnas_extra se reenvía tal cual a scoring_valores.calcular() — ver ahí
    su propósito (solo lo usa el módulo de Test, nunca la importación manual
    de Excel)."""
    respuestas = hojas.get("Respuestas")
    if not respuestas or "Dashboard" in hojas:
        return hojas
    if not scoring_valores.parece_valores_competencias(respuestas):
        return hojas
    scoring_rows, dashboard_rows = scoring_valores.calcular(respuestas, columnas_extra)
    nuevas = dict(hojas)
    nuevas["Scoring"] = scoring_rows
    nuevas["Dashboard"] = dashboard_rows
    return nuevas


def _insertar_hojas(conn, tipo_id, importacion_id, hojas):
    """Inserta cada fila de cada hoja en informe_respuestas, con el mismo
    dedup por hash que usa el import manual de Excel — factorizado aquí para
    que el módulo de Test (encuestas.py) pueda alimentar Informes fila a fila
    según se van recibiendo respuestas, sin duplicar esta lógica.

    ids_por_hoja: id de la fila insertada (o ya existente) en cada hoja —
    para una importación de Excel con muchas filas por hoja solo queda el
    último, pero a quien SÍ le interesa esto es ingest_fila_directa (una
    fila por hoja), para poder enlazar la respuesta del Test directamente
    con su fila en Informes."""
    resumen = {}
    total_nuevas = 0
    ids_por_hoja = {}
    for hoja_nombre, filas in hojas.items():
        nuevas = 0
        ya_existian = 0
        for fila in filas:
            fila_hash = hash_fila(fila)
            existe = conn.execute(
                "SELECT id FROM informe_respuestas WHERE tipo_id = ? AND hoja = ? AND fila_hash = ?",
                (tipo_id, hoja_nombre, fila_hash),
            ).fetchone()
            if existe:
                ya_existian += 1
                ids_por_hoja[hoja_nombre] = existe["id"]
                continue
            datos_json = json.dumps(fila, ensure_ascii=False, default=str)
            cur = conn.execute(
                "INSERT INTO informe_respuestas (tipo_id, importacion_id, hoja, fila_hash, datos_json) "
                "VALUES (?, ?, ?, ?, ?)",
                (tipo_id, importacion_id, hoja_nombre, fila_hash, datos_json),
            )
            ids_por_hoja[hoja_nombre] = cur.lastrowid
            nuevas += 1
        resumen[hoja_nombre] = {"total_en_excel": len(filas), "nuevas": nuevas, "ya_existian": ya_existian}
        total_nuevas += nuevas
    return resumen, total_nuevas, ids_por_hoja


def import_excel(tipo_clave, file_bytes, archivo_nombre, subido_por):
    tipo = get_tipo(tipo_clave)
    if tipo is None:
        raise ValueError(f"Tipo de informe desconocido: {tipo_clave}")

    hojas = read_workbook_sheets(file_bytes)
    if not hojas:
        raise ValueError("El Excel no tiene filas de datos en ninguna hoja")
    hojas = _quiza_calcular_scoring(hojas)

    conn = get_connection()
    cur = conn.execute(
        "INSERT INTO informe_importaciones (tipo_id, archivo_nombre, subido_por, num_respuestas) VALUES (?, ?, ?, 0)",
        (tipo["id"], archivo_nombre, subido_por),
    )
    importacion_id = cur.lastrowid

    resumen, total_nuevas, _ids_por_hoja = _insertar_hojas(conn, tipo["id"], importacion_id, hojas)

    conn.execute("UPDATE informe_importaciones SET num_respuestas = ? WHERE id = ?", (total_nuevas, importacion_id))
    conn.commit()
    conn.close()
    return {"hojas": resumen, "total_nuevas": total_nuevas}


def ingest_fila_directa(tipo_clave, fila, origen="Formulario web", columnas_extra=None):
    """Alimenta Informes con UNA fila recién recibida (p.ej. desde el módulo
    de Test cuando alguien envía el formulario público) exactamente igual
    que si viniera de un Excel: mismo detector de Valores y Competencias,
    mismo cálculo de Scoring/Dashboard, mismo dedup por hash — así el
    resultado aparece en Informes sin duplicar ninguna lógica.

    columnas_extra: preguntas (abiertas / ordenar prioridades) que el
    módulo de Test quiere ver tal cual en Scoring/Dashboard, no solo en la
    hoja "Respuestas" — ver scoring_valores.calcular().

    Devuelve además en qué fila de Informes quedó esta respuesta concreta
    (hoja + id), para que el módulo de Test pueda enlazar directamente a
    "cuál fue la respuesta registrada" en vez de solo el JSON crudo interno."""
    tipo = get_tipo(tipo_clave)
    if tipo is None:
        raise ValueError(f"Tipo de informe desconocido: {tipo_clave}")

    hojas = _quiza_calcular_scoring({"Respuestas": [fila]}, columnas_extra)

    conn = get_connection()
    cur = conn.execute(
        "INSERT INTO informe_importaciones (tipo_id, archivo_nombre, subido_por, num_respuestas) VALUES (?, ?, ?, 0)",
        (tipo["id"], origen, "sistema"),
    )
    importacion_id = cur.lastrowid

    _resumen, total_nuevas, ids_por_hoja = _insertar_hojas(conn, tipo["id"], importacion_id, hojas)

    conn.execute("UPDATE informe_importaciones SET num_respuestas = ? WHERE id = ?", (total_nuevas, importacion_id))
    conn.commit()
    conn.close()

    # Preferimos enlazar a "Dashboard" (vista resumida con puntuación) si el
    # cuestionario se reconoció como Valores y Competencias; si no (p.ej. un
    # tipo de Informe sin scoring), a la propia hoja "Respuestas".
    hoja_destino = "Dashboard" if "Dashboard" in ids_por_hoja else "Respuestas"
    return {
        "total_nuevas": total_nuevas,
        "tipo_clave": tipo_clave,
        "hoja": hoja_destino,
        "respuesta_id": ids_por_hoja.get(hoja_destino),
    }


def forzar_no_apto(respuesta_id):
    """Fuerza RESULTADO = "No apto" en una respuesta ya insertada -- usado
    cuando quien respondió un Test marcó una opción configurada como
    descalificatoria (ver encuestas.guardar_respuesta), sin importar lo que
    scoring_valores.calcular() hubiera calculado. El texto contiene "No
    apto" a propósito, igual que el que genera scoring_valores, para caer
    en el mismo filtro filtro_aptos de get_respuestas() sin duplicar esa
    lógica."""
    conn = get_connection()
    conn.execute(
        "UPDATE informe_respuestas SET datos_json = json_set(datos_json, '$.RESULTADO', ?) WHERE id = ?",
        ("❌ No apto (respuesta descalificatoria)", respuesta_id),
    )
    conn.commit()
    conn.close()


def _detect_date_columns(columnas):
    return [c for c in columnas if any(hint in c.lower() for hint in DATE_HINTS)]


def get_respuestas(tipo_clave, hoja=None, page=1, page_size=200, q=None, orden=None, orden_dir="asc",
                    fecha_col=None, fecha_desde=None, fecha_hasta=None, filtro_aptos="todos"):
    tipo = get_tipo(tipo_clave)
    if tipo is None:
        raise ValueError(f"Tipo de informe desconocido: {tipo_clave}")

    conn = get_connection()
    if hoja is None:
        row = conn.execute("""
            SELECT hoja FROM informe_respuestas
            WHERE tipo_id = ? AND hoja NOT IN (SELECT hoja FROM informe_hojas_ocultas WHERE tipo_id = ?)
            GROUP BY hoja ORDER BY hoja LIMIT 1
        """, (tipo["id"], tipo["id"])).fetchone()
        hoja = row["hoja"] if row else "Scoring"

    clauses = ["tipo_id = ?", "hoja = ?"]
    params = [tipo["id"], hoja]
    if q:
        clauses.append("datos_json LIKE ?")
        params.append(f"%{q}%")
    if fecha_col and fecha_desde:
        clauses.append("json_extract(datos_json, ?) >= ?")
        params.extend([f"$.\"{fecha_col}\"", fecha_desde])
    if fecha_col and fecha_hasta:
        clauses.append("json_extract(datos_json, ?) <= ?")
        params.extend([f"$.\"{fecha_col}\"", fecha_hasta])
    if filtro_aptos == "excluir_no_aptos":
        # RESULTADO viene del Excel de Valores y Competencias como "❌ No
        # apto" — otros tipos de informe no tienen esta columna, así que se
        # deja pasar cuando es NULL en vez de excluir de más.
        clauses.append("(json_extract(datos_json, '$.RESULTADO') IS NULL OR json_extract(datos_json, '$.RESULTADO') NOT LIKE '%No apto%')")
    elif filtro_aptos == "solo_no_aptos":
        clauses.append("json_extract(datos_json, '$.RESULTADO') LIKE '%No apto%'")
    where = "WHERE " + " AND ".join(clauses)

    order_sql = "id DESC"
    if orden:
        direction = "ASC" if orden_dir == "asc" else "DESC"
        order_sql = f'json_extract(datos_json, \'$."{orden}"\') {direction}'

    total = conn.execute(f"SELECT COUNT(*) FROM informe_respuestas {where}", params).fetchone()[0]
    offset = (page - 1) * page_size
    rows = conn.execute(
        f"SELECT id, datos_json, cv_ruta, cv_nombre_original, creado_en FROM informe_respuestas {where} "
        f"ORDER BY {order_sql} LIMIT ? OFFSET ?",
        params + [page_size, offset],
    ).fetchall()

    ids_pagina = [row["id"] for row in rows]
    compartidos_por_respuesta = {}
    if ids_pagina:
        placeholders = ",".join("?" * len(ids_pagina))
        # Unión de los dos caminos de compartir, igual que
        # reclutamiento._compartidos_por_candidato -- antes esto solo miraba
        # informe_compartidos y se dejaba fuera a quien se hubiera compartido
        # como "directo" desde Reclutamiento (candidato_compartidos), así que
        # el sombreado verde y el aviso de "ya compartido" de Informes iban
        # por debajo de lo que de verdad había (se veían menos filas
        # compartidas aquí que en Reclutamiento, que sí unía las dos).
        for r in conn.execute(f"""
            SELECT respuesta_id, usuario_nombre FROM (
                SELECT ic.respuesta_id AS respuesta_id, u.nombre AS usuario_nombre
                FROM informe_compartidos ic JOIN usuarios u ON u.id = ic.usuario_id
                WHERE ic.respuesta_id IN ({placeholders})
                UNION
                SELECT cand.respuesta_id AS respuesta_id, u.nombre AS usuario_nombre
                FROM candidatos cand
                JOIN candidato_compartidos cc ON cc.candidato_id = cand.id
                JOIN usuarios u ON u.id = cc.usuario_id
                WHERE cand.respuesta_id IN ({placeholders})
            )
        """, ids_pagina + ids_pagina).fetchall():
            compartidos_por_respuesta.setdefault(r["respuesta_id"], []).append(r["usuario_nombre"])
    conn.close()

    respuestas = []
    columnas = []
    seen = set()
    for row in rows:
        datos = json.loads(row["datos_json"])
        for k in datos.keys():
            if k not in seen:
                seen.add(k)
                columnas.append(k)
        respuestas.append({
            "id": row["id"],
            "creado_en": row["creado_en"],
            "datos": datos,
            "tiene_cv": row["cv_ruta"] is not None,
            "cv_nombre": row["cv_nombre_original"],
            "compartido_con": compartidos_por_respuesta.get(row["id"], []),
        })

    return {
        "tipo": tipo,
        "hoja": hoja,
        "total": total,
        "columnas": columnas,
        "columnas_fecha": _detect_date_columns(columnas),
        "respuestas": respuestas,
    }


def get_respuesta(respuesta_id):
    conn = get_connection()
    row = conn.execute("SELECT * FROM informe_respuestas WHERE id = ?", (respuesta_id,)).fetchone()
    conn.close()
    return dict(row) if row else None


def guardar_cv(respuesta_id, archivo_nombre, contenido):
    respuesta = get_respuesta(respuesta_id)
    if respuesta is None:
        raise ValueError("Respuesta no encontrada")
    os.makedirs(CV_DIR, exist_ok=True)
    ext = os.path.splitext(archivo_nombre)[1]
    ruta = os.path.join(CV_DIR, f"{respuesta_id}{ext}")
    with open(ruta, "wb") as f:
        f.write(contenido)
    conn = get_connection()
    conn.execute(
        "UPDATE informe_respuestas SET cv_ruta = ?, cv_nombre_original = ? WHERE id = ?",
        (ruta, archivo_nombre, respuesta_id),
    )
    conn.commit()
    conn.close()


def _candidato_id_para_respuesta(respuesta_id, compartido_por):
    """Devuelve el candidato ya asociado a esta respuesta o crea uno nuevo
    (origen='informe'), mapeando datos_json a los campos conocidos del
    candidato (ver reclutamiento.mapear_datos_a_candidato) y enlazando el CV
    ya subido, si lo hay. Así "compartir" deja una ficha real y editable en
    vez de una vista de solo lectura del Excel."""
    existente = reclutamiento_module.get_candidato_por_respuesta(respuesta_id)
    if existente is not None:
        return existente

    respuesta = get_respuesta(respuesta_id)
    if respuesta is None:
        raise ValueError("Respuesta no encontrada")

    conn = get_connection()
    tipo_row = conn.execute("SELECT empresa FROM informe_tipos WHERE id = ?", (respuesta["tipo_id"],)).fetchone()
    conn.close()
    empresa = tipo_row["empresa"] if tipo_row else "kk"

    datos = json.loads(respuesta["datos_json"])
    campos, extra = reclutamiento_module.mapear_datos_a_candidato(datos)

    # Si esta persona ya está en Reclutamiento (creada a mano, por CV, o por
    # una vacante) pero todavía sin ningún test enlazado, se conecta ahí en
    # vez de crear una ficha duplicada — mismo criterio que el match
    # automático al recibir la respuesta (ver encuestas.guardar_respuesta).
    candidato_existente = reclutamiento_module.buscar_candidato_sin_respuesta_por_contacto(
        campos.get("telefono"), campos.get("email")
    )
    if candidato_existente:
        reclutamiento_module.enlazar_respuesta_a_candidato(candidato_existente, respuesta_id)
        return candidato_existente

    campos["extra_fields"] = extra
    candidato_id = reclutamiento_module.crear_candidato(
        campos, empresa=empresa, origen="informe", respuesta_id=respuesta_id, creado_por=compartido_por
    )
    if respuesta["cv_ruta"]:
        reclutamiento_module.registrar_archivo_existente(
            candidato_id, respuesta["cv_nombre_original"] or "cv", respuesta["cv_ruta"]
        )
    return candidato_id


def compartir_respuestas(respuesta_ids, usuario_id, compartido_por):
    """Comparte cada respuesta con `usuario_id` -- igual que
    reclutamiento.compartir_candidatos_directo, esto es EXCLUSIVO: si el
    candidato ligado a la respuesta ya estaba compartido con otra persona
    (por este camino o por el directo de Reclutamiento), se le quita el
    acceso a esa persona antes de dárselo al nuevo destinatario."""
    conn = get_connection()
    for rid in respuesta_ids:
        candidato_id = _candidato_id_para_respuesta(rid, compartido_por)
        conn.execute(
            "DELETE FROM candidato_compartidos WHERE candidato_id = ? AND usuario_id != ?",
            (candidato_id, usuario_id),
        )
        conn.execute(
            "DELETE FROM informe_compartidos WHERE candidato_id = ? AND usuario_id != ?",
            (candidato_id, usuario_id),
        )
        # Upsert en vez de INSERT OR IGNORE: si ya se había compartido antes,
        # volver a compartir debe refrescar la fecha (y quién lo hizo), para
        # que el orden de Reclutamiento refleje la última vez que se compartió.
        conn.execute(
            """
            INSERT INTO informe_compartidos (respuesta_id, usuario_id, compartido_por, candidato_id)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(respuesta_id, usuario_id)
            DO UPDATE SET compartido_por = excluded.compartido_por, compartido_en = datetime('now'),
                          candidato_id = excluded.candidato_id
            """,
            (rid, usuario_id, compartido_por, candidato_id),
        )
    conn.commit()
    conn.close()


def dejar_de_compartir(respuesta_id, usuario_id):
    conn = get_connection()
    conn.execute(
        "DELETE FROM informe_compartidos WHERE respuesta_id = ? AND usuario_id = ?", (respuesta_id, usuario_id)
    )
    conn.commit()
    conn.close()


def cambiar_destinatario_informe(pares, nuevo_usuario_id, compartido_en):
    """Igual que reclutamiento.cambiar_destinatario_directo pero para
    informe_compartidos (pares = [(respuesta_id, usuario_id_actual), ...])."""
    conn = get_connection()
    for respuesta_id, usuario_id_actual in pares:
        if usuario_id_actual == nuevo_usuario_id:
            conn.execute(
                "UPDATE informe_compartidos SET compartido_en = ? WHERE respuesta_id = ? AND usuario_id = ?",
                (compartido_en, respuesta_id, nuevo_usuario_id),
            )
            continue
        ya_existe = conn.execute(
            "SELECT 1 FROM informe_compartidos WHERE respuesta_id = ? AND usuario_id = ?",
            (respuesta_id, nuevo_usuario_id),
        ).fetchone()
        if ya_existe:
            conn.execute(
                "DELETE FROM informe_compartidos WHERE respuesta_id = ? AND usuario_id = ?",
                (respuesta_id, usuario_id_actual),
            )
            conn.execute(
                "UPDATE informe_compartidos SET compartido_en = ? WHERE respuesta_id = ? AND usuario_id = ?",
                (compartido_en, respuesta_id, nuevo_usuario_id),
            )
        else:
            conn.execute(
                "UPDATE informe_compartidos SET usuario_id = ?, compartido_en = ? "
                "WHERE respuesta_id = ? AND usuario_id = ?",
                (nuevo_usuario_id, compartido_en, respuesta_id, usuario_id_actual),
            )
    conn.commit()
    conn.close()


def cambiar_destinatario_compartidos(items, nuevo_usuario_id):
    """Orquesta el cambio de destinatario/fusión de tanda para una
    selección mixta de "Compartidos por ti" (candidatos compartidos directo
    + candidatos que llegaron vía Informes -- ver frontend/js/compartidos.js).
    `items` son dicts {tipo: 'directo'|'informe', candidato_id o
    respuesta_id, usuario_id_actual}. Todos quedan re-estampados con el
    MISMO timestamp (calculado una sola vez aquí), así que sea cual sea su
    tanda/destinatario original, aparecen agrupados juntos después -- esto
    cubre a la vez "cambiar a quién se compartió" (Heber -> Adhara) y
    "fusionar tandas sueltas en una sola" (mismo destinatario, timestamps
    distintos)."""
    ahora = datetime.datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
    pares_directo = [(it["candidato_id"], it["usuario_id_actual"]) for it in items if it["tipo"] == "directo"]
    pares_informe = [(it["respuesta_id"], it["usuario_id_actual"]) for it in items if it["tipo"] == "informe"]
    if pares_directo:
        reclutamiento_module.cambiar_destinatario_directo(pares_directo, nuevo_usuario_id, ahora)
    if pares_informe:
        cambiar_destinatario_informe(pares_informe, nuevo_usuario_id, ahora)


def usuario_tiene_acceso_respuesta(usuario_id, respuesta_id):
    conn = get_connection()
    row = conn.execute(
        "SELECT 1 FROM informe_compartidos WHERE respuesta_id = ? AND usuario_id = ?", (respuesta_id, usuario_id)
    ).fetchone()
    conn.close()
    return row is not None


def get_tipos_permitidos(usuario_id: int) -> list[str]:
    """Claves de tipo de informe a las que este usuario tiene acceso.
    Lista vacía = sin restricción (ve todos, como usuario_tiendas)."""
    conn = get_connection()
    rows = conn.execute(
        "SELECT tipo_clave FROM usuario_informe_tipos WHERE usuario_id = ? ORDER BY tipo_clave", (usuario_id,)
    ).fetchall()
    conn.close()
    return [r["tipo_clave"] for r in rows]


def set_tipos_permitidos(usuario_id: int, tipos: list[str]):
    conn = get_connection()
    conn.execute("DELETE FROM usuario_informe_tipos WHERE usuario_id = ?", (usuario_id,))
    for tipo_clave in tipos:
        conn.execute(
            "INSERT OR IGNORE INTO usuario_informe_tipos (usuario_id, tipo_clave) VALUES (?, ?)", (usuario_id, tipo_clave)
        )
    conn.commit()
    conn.close()


def _candidato_directo_como_item(row, incluir_destinatario=False):
    """Adapta una fila de candidato_compartidos (sin respuesta de Informes
    detrás) a la misma forma que devuelve get_compartidos_con/por, para que
    la pantalla de "Compartidos" pueda pintar ambos orígenes con el mismo
    candidatoCardHTML sin ramificar el frontend. `datos` se sintetiza a
    partir de los campos del candidato en vez de venir de un Excel."""
    datos = {}
    if row["nombre_completo"]:
        datos["Nombre"] = row["nombre_completo"]
    if row["telefono"]:
        datos["Teléfono"] = row["telefono"]
    if row["email"]:
        datos["Email"] = row["email"]
    if row["puesto_solicitado"]:
        datos["Puesto solicitado"] = row["puesto_solicitado"]
    item = {
        "compartido_id": f"candidato-{row['compartido_id']}",
        "compartido_en": row["compartido_en"],
        "compartido_por": row["compartido_por"],
        "respuesta_id": row["respuesta_id"],
        "datos": datos,
        "hoja": None,
        "tiene_cv": False,
        "cv_nombre": None,
        "tipo_nombre": "Candidato (Reclutamiento)",
        "tipo_clave": None,
        "candidato_id": row["id"],
        "estado": row["estado"],
        "notas": row["notas"],
        "telefono": row["telefono"],
        "puesto_solicitado": row["puesto_solicitado"],
        "test_resultado": row["test_resultado"],
        "vacante_id": row["vacante_id"],
    }
    if incluir_destinatario:
        item["destinatario_id"] = row["destinatario_id"]
        item["destinatario_nombre"] = row["destinatario_nombre"]
        item["destinatario_username"] = row["destinatario_username"]
    return item


def get_compartidos_con(usuario_id, empresa=None):
    conn = get_connection()
    clauses = ["c.usuario_id = ?"]
    params = [usuario_id]
    if empresa:
        clauses.append("t.empresa = ?")
        params.append(empresa)
    rows = conn.execute(f"""
        SELECT c.id AS compartido_id, c.compartido_en, c.compartido_por, c.candidato_id,
               r.id AS respuesta_id, r.datos_json, r.hoja, r.cv_ruta, r.cv_nombre_original,
               t.nombre AS tipo_nombre, t.clave AS tipo_clave,
               cand.estado AS candidato_estado, cand.notas AS candidato_notas,
               cand.telefono AS candidato_telefono, cand.puesto_solicitado AS candidato_puesto
        FROM informe_compartidos c
        JOIN informe_respuestas r ON r.id = c.respuesta_id
        JOIN informe_tipos t ON t.id = r.tipo_id
        LEFT JOIN candidatos cand ON cand.id = c.candidato_id
        WHERE {' AND '.join(clauses)}
        ORDER BY c.compartido_en DESC
    """, params).fetchall()
    conn.close()
    resultado = []
    for row in rows:
        datos = json.loads(row["datos_json"])
        resultado.append({
            "compartido_id": row["compartido_id"],
            "compartido_en": row["compartido_en"],
            "compartido_por": row["compartido_por"],
            "respuesta_id": row["respuesta_id"],
            "datos": datos,
            "hoja": row["hoja"],
            "tiene_cv": row["cv_ruta"] is not None,
            "cv_nombre": row["cv_nombre_original"],
            "tipo_nombre": row["tipo_nombre"],
            "tipo_clave": row["tipo_clave"],
            "candidato_id": row["candidato_id"],
            "estado": row["candidato_estado"],
            "notas": row["candidato_notas"],
            "telefono": row["candidato_telefono"],
            "puesto_solicitado": row["candidato_puesto"],
            "test_resultado": datos.get("RESULTADO"),
        })
    directos = reclutamiento_module.get_candidatos_compartidos_directo_con(usuario_id, empresa=empresa)
    resultado += [_candidato_directo_como_item(r) for r in directos]
    resultado.sort(key=lambda it: it["compartido_en"], reverse=True)
    return resultado


def get_compartidos_por(username, empresa=None):
    """Candidatos que ESTE usuario ha compartido con otros (para su propia
    carpeta de Reclutamiento, sección "Compartidos por ti"). Incluye a quién
    se lo compartió, para poder agrupar por tanda + destinatario."""
    conn = get_connection()
    clauses = ["c.compartido_por = ?"]
    params = [username]
    if empresa:
        clauses.append("t.empresa = ?")
        params.append(empresa)
    rows = conn.execute(f"""
        SELECT c.id AS compartido_id, c.compartido_en, c.compartido_por, c.candidato_id,
               c.usuario_id AS destinatario_id,
               u.nombre AS destinatario_nombre, u.username AS destinatario_username,
               r.id AS respuesta_id, r.datos_json, r.hoja, r.cv_ruta, r.cv_nombre_original,
               t.nombre AS tipo_nombre, t.clave AS tipo_clave,
               cand.estado AS candidato_estado, cand.notas AS candidato_notas,
               cand.telefono AS candidato_telefono, cand.puesto_solicitado AS candidato_puesto,
               cand.vacante_id AS candidato_vacante_id
        FROM informe_compartidos c
        JOIN informe_respuestas r ON r.id = c.respuesta_id
        JOIN informe_tipos t ON t.id = r.tipo_id
        JOIN usuarios u ON u.id = c.usuario_id
        LEFT JOIN candidatos cand ON cand.id = c.candidato_id
        WHERE {' AND '.join(clauses)}
        ORDER BY c.compartido_en DESC
    """, params).fetchall()
    conn.close()
    resultado = []
    for row in rows:
        datos = json.loads(row["datos_json"])
        resultado.append({
            "compartido_id": row["compartido_id"],
            "compartido_en": row["compartido_en"],
            "destinatario_id": row["destinatario_id"],
            "destinatario_nombre": row["destinatario_nombre"],
            "destinatario_username": row["destinatario_username"],
            "respuesta_id": row["respuesta_id"],
            "datos": datos,
            "hoja": row["hoja"],
            "vacante_id": row["candidato_vacante_id"],
            "tiene_cv": row["cv_ruta"] is not None,
            "cv_nombre": row["cv_nombre_original"],
            "tipo_nombre": row["tipo_nombre"],
            "tipo_clave": row["tipo_clave"],
            "candidato_id": row["candidato_id"],
            "estado": row["candidato_estado"],
            "notas": row["candidato_notas"],
            "telefono": row["candidato_telefono"],
            "puesto_solicitado": row["candidato_puesto"],
            "test_resultado": datos.get("RESULTADO"),
        })
    directos = reclutamiento_module.get_candidatos_compartidos_directo_por(username, empresa=empresa)
    resultado += [_candidato_directo_como_item(r, incluir_destinatario=True) for r in directos]
    resultado.sort(key=lambda it: it["compartido_en"], reverse=True)
    return resultado


ensure_informe_tables()
