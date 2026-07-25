"""Boletines/blog interno: publicaciones que se ven en una página pública
(sin login) y que además se pueden mandar por email a una lista de
contactos elegida a mano en cada envío (no hay suscripción automática — el
propio usuario decide a quién le llega cada boletín)."""
import base64
import io
import os

import requests
from openpyxl import load_workbook

from db import get_connection

RESEND_API_URL = "https://api.resend.com/emails"
PDF_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "uploads", "boletines"))
IMAGENES_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "uploads", "boletines_imagenes"))


def ensure_boletin_tables():
    conn = get_connection()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS boletin_posts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            titulo TEXT NOT NULL,
            resumen TEXT,
            contenido_html TEXT NOT NULL,
            publicado INTEGER NOT NULL DEFAULT 0,
            autor TEXT,
            creado_en TEXT NOT NULL DEFAULT (datetime('now')),
            publicado_en TEXT
        )
    """)
    cols_posts = {row[1] for row in conn.execute("PRAGMA table_info(boletin_posts)")}
    if "pdf_ruta" not in cols_posts:
        conn.execute("ALTER TABLE boletin_posts ADD COLUMN pdf_ruta TEXT")
    if "pdf_nombre_original" not in cols_posts:
        conn.execute("ALTER TABLE boletin_posts ADD COLUMN pdf_nombre_original TEXT")
    if "contenido_bloques" not in cols_posts:
        conn.execute("ALTER TABLE boletin_posts ADD COLUMN contenido_bloques TEXT")
    os.makedirs(PDF_DIR, exist_ok=True)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS boletin_imagenes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            post_id INTEGER NOT NULL REFERENCES boletin_posts(id),
            ruta TEXT NOT NULL,
            nombre_original TEXT,
            subido_en TEXT NOT NULL DEFAULT (datetime('now'))
        )
    """)
    os.makedirs(IMAGENES_DIR, exist_ok=True)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS boletin_contactos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nombre TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            activo INTEGER NOT NULL DEFAULT 1,
            creado_en TEXT NOT NULL DEFAULT (datetime('now'))
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS boletin_envios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            post_id INTEGER NOT NULL REFERENCES boletin_posts(id),
            contacto_id INTEGER NOT NULL REFERENCES boletin_contactos(id),
            enviado_en TEXT NOT NULL DEFAULT (datetime('now')),
            ok INTEGER NOT NULL,
            error TEXT
        )
    """)
    conn.commit()
    conn.close()


def _row_post(r):
    d = dict(r)
    d["publicado"] = bool(d["publicado"])
    d["tiene_pdf"] = d["pdf_ruta"] is not None
    del d["pdf_ruta"]
    return d


def list_posts(solo_publicados=False):
    conn = get_connection()
    if solo_publicados:
        rows = conn.execute(
            "SELECT * FROM boletin_posts WHERE publicado = 1 ORDER BY publicado_en DESC"
        ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM boletin_posts ORDER BY id DESC").fetchall()
    conn.close()
    return [_row_post(r) for r in rows]


def get_post(post_id, solo_publicado=False):
    conn = get_connection()
    row = conn.execute("SELECT * FROM boletin_posts WHERE id = ?", (post_id,)).fetchone()
    conn.close()
    if not row:
        return None
    post = _row_post(row)
    if solo_publicado and not post["publicado"]:
        return None
    return post


def create_post(titulo, resumen, contenido_html, autor, contenido_bloques=None):
    conn = get_connection()
    cur = conn.execute(
        "INSERT INTO boletin_posts (titulo, resumen, contenido_html, autor, contenido_bloques) VALUES (?, ?, ?, ?, ?)",
        (titulo, resumen, contenido_html, autor, contenido_bloques),
    )
    conn.commit()
    post_id = cur.lastrowid
    conn.close()
    return post_id


def update_post(post_id, titulo, resumen, contenido_html, contenido_bloques=None):
    conn = get_connection()
    conn.execute(
        "UPDATE boletin_posts SET titulo = ?, resumen = ?, contenido_html = ?, contenido_bloques = ? WHERE id = ?",
        (titulo, resumen, contenido_html, contenido_bloques, post_id),
    )
    conn.commit()
    conn.close()


def publicar_post(post_id):
    conn = get_connection()
    conn.execute(
        "UPDATE boletin_posts SET publicado = 1, publicado_en = datetime('now') WHERE id = ? AND publicado = 0",
        (post_id,),
    )
    conn.commit()
    conn.close()


def despublicar_post(post_id):
    conn = get_connection()
    conn.execute("UPDATE boletin_posts SET publicado = 0, publicado_en = NULL WHERE id = ?", (post_id,))
    conn.commit()
    conn.close()


def delete_post(post_id):
    _borrar_pdf_fisico(post_id)
    for imagen in list_imagenes(post_id):
        _borrar_imagen_fisica(imagen["ruta"])
    conn = get_connection()
    conn.execute("DELETE FROM boletin_envios WHERE post_id = ?", (post_id,))
    conn.execute("DELETE FROM boletin_imagenes WHERE post_id = ?", (post_id,))
    conn.execute("DELETE FROM boletin_posts WHERE id = ?", (post_id,))
    conn.commit()
    conn.close()


def _borrar_pdf_fisico(post_id):
    conn = get_connection()
    row = conn.execute("SELECT pdf_ruta FROM boletin_posts WHERE id = ?", (post_id,)).fetchone()
    conn.close()
    if row and row["pdf_ruta"] and os.path.exists(row["pdf_ruta"]):
        os.remove(row["pdf_ruta"])


def get_pdf_info(post_id, solo_publicado=False):
    """Devuelve la ruta física + nombre original del PDF, para servir el
    archivo — separado de get_post porque ese ya no expone la ruta real."""
    conn = get_connection()
    row = conn.execute("SELECT publicado, pdf_ruta, pdf_nombre_original FROM boletin_posts WHERE id = ?", (post_id,)).fetchone()
    conn.close()
    if not row or not row["pdf_ruta"]:
        return None
    if solo_publicado and not row["publicado"]:
        return None
    return {"ruta": row["pdf_ruta"], "nombre": row["pdf_nombre_original"] or "boletin.pdf"}


def guardar_pdf(post_id, nombre_original, contenido):
    if not get_post(post_id):
        raise ValueError("El boletín no existe")
    _borrar_pdf_fisico(post_id)
    ruta = os.path.join(PDF_DIR, f"{post_id}.pdf")
    with open(ruta, "wb") as f:
        f.write(contenido)
    conn = get_connection()
    conn.execute(
        "UPDATE boletin_posts SET pdf_ruta = ?, pdf_nombre_original = ? WHERE id = ?",
        (ruta, nombre_original, post_id),
    )
    conn.commit()
    conn.close()


def _borrar_imagen_fisica(ruta):
    if ruta and os.path.exists(ruta):
        os.remove(ruta)


def guardar_imagen(post_id, nombre_original, extension, contenido):
    """Guarda una imagen de un bloque del builder (foto, encabezado...). Un
    boletín puede tener varias, a diferencia del único PDF adjunto, así que
    cada una vive en su propia fila de boletin_imagenes en vez de una
    columna del post."""
    if not get_post(post_id):
        raise ValueError("El boletín no existe")
    conn = get_connection()
    cur = conn.execute(
        "INSERT INTO boletin_imagenes (post_id, ruta, nombre_original) VALUES (?, ?, ?)",
        (post_id, "", nombre_original),
    )
    imagen_id = cur.lastrowid
    ruta = os.path.join(IMAGENES_DIR, f"{imagen_id}{extension}")
    with open(ruta, "wb") as f:
        f.write(contenido)
    conn.execute("UPDATE boletin_imagenes SET ruta = ? WHERE id = ?", (ruta, imagen_id))
    conn.commit()
    conn.close()
    return imagen_id


def list_imagenes(post_id):
    conn = get_connection()
    rows = conn.execute(
        "SELECT * FROM boletin_imagenes WHERE post_id = ? ORDER BY id", (post_id,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def get_imagen_info(post_id, imagen_id):
    conn = get_connection()
    row = conn.execute(
        "SELECT * FROM boletin_imagenes WHERE id = ? AND post_id = ?", (imagen_id, post_id)
    ).fetchone()
    conn.close()
    return dict(row) if row else None


def borrar_imagen(post_id, imagen_id):
    imagen = get_imagen_info(post_id, imagen_id)
    if not imagen:
        return
    _borrar_imagen_fisica(imagen["ruta"])
    conn = get_connection()
    conn.execute("DELETE FROM boletin_imagenes WHERE id = ?", (imagen_id,))
    conn.commit()
    conn.close()


def list_contactos(solo_activos=False):
    conn = get_connection()
    if solo_activos:
        rows = conn.execute("SELECT * FROM boletin_contactos WHERE activo = 1 ORDER BY nombre").fetchall()
    else:
        rows = conn.execute("SELECT * FROM boletin_contactos ORDER BY nombre").fetchall()
    conn.close()
    return [dict(r) for r in rows]


def add_contacto(nombre, email):
    conn = get_connection()
    try:
        cur = conn.execute("INSERT INTO boletin_contactos (nombre, email) VALUES (?, ?)", (nombre, email))
        conn.commit()
        return cur.lastrowid
    finally:
        conn.close()


def delete_contacto(contacto_id):
    conn = get_connection()
    conn.execute("DELETE FROM boletin_contactos WHERE id = ?", (contacto_id,))
    conn.commit()
    conn.close()


def _normaliza_header(h):
    return (h or "").strip().lower().replace("í", "i").replace("ó", "o").replace("á", "a").replace(
        "é", "e"
    ).replace("ú", "u")


def import_contactos_excel(file_bytes):
    """Sube en bloque una lista de contactos desde un Excel con columnas de
    nombre y email (el nombre exacto de las columnas es flexible: admite
    "Nombre"/"Name", "Email"/"Correo"/"E-mail"...). Los emails repetidos se
    ignoran en vez de fallar toda la importación."""
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    except Exception as exc:
        raise ValueError(f"No se pudo leer el archivo Excel: {exc}")

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = [str(h).strip() if h is not None else "" for h in next(rows_iter)]
    except StopIteration:
        raise ValueError("El Excel no tiene filas")

    idx_nombre = idx_email = None
    for i, h in enumerate(header):
        norm = _normaliza_header(h)
        if idx_nombre is None and ("nombre" in norm or norm == "name"):
            idx_nombre = i
        if idx_email is None and ("email" in norm or "correo" in norm or "e-mail" in norm):
            idx_email = i
    if idx_email is None:
        raise ValueError("No se encontró una columna de email en el Excel")

    conn = get_connection()
    nuevos = 0
    ya_existian = 0
    invalidos = 0
    for row in rows_iter:
        if row is None or all(v is None for v in row):
            continue
        email = str(row[idx_email]).strip().lower() if idx_email < len(row) and row[idx_email] else ""
        nombre = str(row[idx_nombre]).strip() if idx_nombre is not None and idx_nombre < len(row) and row[idx_nombre] else email
        if "@" not in email:
            invalidos += 1
            continue
        existe = conn.execute("SELECT id FROM boletin_contactos WHERE email = ?", (email,)).fetchone()
        if existe:
            ya_existian += 1
            continue
        conn.execute("INSERT INTO boletin_contactos (nombre, email) VALUES (?, ?)", (nombre, email))
        nuevos += 1
    conn.commit()
    conn.close()
    return {"nuevos": nuevos, "ya_existian": ya_existian, "invalidos": invalidos}


def _enviar_email_resend(destino_email, destino_nombre, asunto, html, adjunto=None):
    api_key = os.environ.get("RESEND_API_KEY")
    remitente = os.environ.get("BOLETIN_FROM_EMAIL", "onboarding@resend.dev")
    if not api_key:
        return False, "Falta configurar RESEND_API_KEY en el entorno del servidor"
    payload = {
        "from": remitente,
        "to": [destino_email],
        "subject": asunto,
        "html": html,
    }
    if adjunto:
        payload["attachments"] = [{"filename": adjunto["nombre"], "content": adjunto["base64"]}]
    try:
        resp = requests.post(
            RESEND_API_URL,
            headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            json=payload,
            timeout=30,
        )
    except requests.RequestException as exc:
        return False, str(exc)
    if resp.status_code >= 300:
        return False, f"Resend respondió {resp.status_code}: {resp.text[:300]}"
    return True, None


def enviar_boletin(post_id, contacto_ids):
    """Manda el post por email a los contactos indicados y deja constancia
    de cada envío (para no perder el rastro de a quién le llegó ya, aunque
    el usuario decida re-enviar el mismo boletín otro día a gente distinta).
    Si el boletín tiene un PDF adjunto, se manda igual como adjunto del email."""
    post = get_post(post_id)
    if not post:
        raise ValueError("El boletín no existe")

    adjunto = None
    pdf_info = get_pdf_info(post_id)
    if pdf_info and os.path.exists(pdf_info["ruta"]):
        with open(pdf_info["ruta"], "rb") as f:
            adjunto = {"nombre": pdf_info["nombre"], "base64": base64.b64encode(f.read()).decode("ascii")}

    conn = get_connection()
    placeholders = ",".join("?" for _ in contacto_ids)
    rows = conn.execute(
        f"SELECT * FROM boletin_contactos WHERE id IN ({placeholders})", contacto_ids
    ).fetchall() if contacto_ids else []
    conn.close()

    enviados = 0
    fallidos = []
    conn = get_connection()
    for contacto in rows:
        ok, error = _enviar_email_resend(contacto["email"], contacto["nombre"], post["titulo"], post["contenido_html"], adjunto)
        conn.execute(
            "INSERT INTO boletin_envios (post_id, contacto_id, ok, error) VALUES (?, ?, ?, ?)",
            (post_id, contacto["id"], 1 if ok else 0, error),
        )
        if ok:
            enviados += 1
        else:
            fallidos.append({"nombre": contacto["nombre"], "email": contacto["email"], "error": error})
    conn.commit()
    conn.close()
    return {"enviados": enviados, "fallidos": fallidos, "total": len(rows)}


def list_envios(post_id):
    conn = get_connection()
    rows = conn.execute("""
        SELECT e.id, e.enviado_en, e.ok, e.error, c.nombre, c.email
        FROM boletin_envios e JOIN boletin_contactos c ON c.id = e.contacto_id
        WHERE e.post_id = ? ORDER BY e.enviado_en DESC
    """, (post_id,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]


ensure_boletin_tables()
